# ==================== VISTAS DE PRODUCTOS E INGREDIENTES ====================
"""
Módulo de vistas para la gestión de productos e ingredientes (materias primas).
Funcionalidades:
- CRUD de productos con soporte para recetas y fraccionamientos
- CRUD de materias primas
- Movimientos de inventario con FIFO
- Exportación de datos
- Reportes de stock y costos
"""

# ==================== IMPORTS ====================
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import models, transaction
from django.db.models import F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from .forms import (
    MateriaPrimaForm,
    MovimientoMateriaPrimaForm,
    ProductoForm,
)
from .models import (
    LoteMateriaPrima,
    MateriaPrima,
    MovimientoMateriaPrima,
    Producto,
    ProductoMateriaPrima,
    VentaDetalle,
)
from .resources import ProductoResource

# ==================== VISTAS DE PRODUCTOS ====================

@login_required
def lista_productos(request):
    """Vista mejorada de lista de productos con filtros y KPIs LINO V3."""
    from django.core.paginator import Paginator

    from gestion.utils.kpi_builder import prepare_product_kpis

    productos = Producto.objects.all()

    # Filtros
    query = request.GET.get('q', '').strip()
    categoria_seleccionada = request.GET.get('categoria', '')
    estado_stock = request.GET.get('estado_stock', '')

    # Aplicar filtros
    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) |
            Q(descripcion__icontains=query) |
            Q(marca__icontains=query)
        )

    if categoria_seleccionada:
        productos = productos.filter(categoria=categoria_seleccionada)

    if estado_stock:
        if estado_stock == 'agotado':
            productos = productos.filter(stock=0)
        elif estado_stock == 'critico':
            productos = productos.filter(stock__gt=0, stock__lte=F('stock_minimo'))
        elif estado_stock == 'bajo':
            productos = productos.filter(stock__gt=F('stock_minimo'), stock__lte=F('stock_minimo') * 2)
        elif estado_stock == 'normal':
            productos = productos.filter(stock__gt=F('stock_minimo') * 2)

    # Paginación
    paginator = Paginator(productos.order_by('nombre'), 25)
    page_number = request.GET.get('page', 1)
    productos_paginados = paginator.get_page(page_number)

    # Preparar KPIs usando utility
    kpis = prepare_product_kpis(Producto.objects.all())

    # Obtener categorías disponibles para el filtro
    categorias = [choice[0] for choice in Producto.CATEGORIAS_DIETETICA]

    context = {
        'productos': productos_paginados,
        'kpis': kpis,
        'categorias': categorias,
        'query': query,
        'categoria_seleccionada': categoria_seleccionada,
        'estado_stock': estado_stock,
        'title': 'Productos',
        'subtitle': 'Gestión completa del catálogo de productos',
        'icon': 'box-seam',
        'create_url': reverse('gestion:crear_producto'),
        'export_url': reverse('gestion:exportar_productos'),
    }

    return render(request, 'modules/productos/lista.html', context)


@login_required
def crear_producto(request):
    """Vista para crear un nuevo producto."""
    if not request.user.has_perm('gestion.add_producto'):
        messages.error(request, 'No tienes permiso para crear productos.')
        return redirect('gestion:lista_productos')

    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Verificar si se creó una nueva categoría
                    nueva_categoria = form.cleaned_data.get('nueva_categoria')
                    if nueva_categoria and form.cleaned_data.get('categoria') != 'nueva':
                        # La nueva categoría ya fue procesada en el clean() del formulario
                        pass

                    # Guardar el producto primero (sin stock inicial)
                    producto = form.save(commit=False)
                    stock_inicial = form.cleaned_data.get('stock', 0)
                    producto.stock = 0  # Inicializar en 0 temporalmente
                    producto.save()

                    # Ahora que el producto tiene ID, podemos trabajar con las relaciones
                    # Si el producto usa receta, verificar materias primas y producir
                    if stock_inicial and stock_inicial > 0:
                        if producto.tipo_producto == 'receta' and producto.receta:
                            # Verificar stock de materias primas
                            ok, faltantes = producto.verificar_stock_materias_primas(stock_inicial)
                            if not ok:
                                faltantes_str = ", ".join([f"{f['materia_prima']} (necesaria: {f['necesaria']} {f['unidad']}, disponible: {f['disponible']})" for f in faltantes])
                                messages.error(request, f"No hay suficiente stock de materias primas para producir {stock_inicial} unidades: {faltantes_str}")
                                # Eliminar el producto recién creado
                                producto.delete()
                                raise Exception("Stock de materias primas insuficiente")

                            # Descontar materias primas y actualizar stock
                            producto.descontar_materias_primas(stock_inicial, request.user)
                            producto.stock = stock_inicial
                            producto.save()
                        else:
                            # Para productos que no usan receta, simplemente establecer el stock
                            producto.stock = stock_inicial
                            producto.save()

                    # Calcular y actualizar costos después de crear el producto
                    if producto.tipo_producto in ['receta', 'fraccionamiento']:
                        # Calcular costo base automáticamente
                        costo_calculado = producto.calcular_costo_unitario()
                        if costo_calculado > 0:
                            producto.costo_base = costo_calculado

                            # Calcular precio sugerido si hay margen
                            if producto.margen_ganancia and producto.margen_ganancia > 0:
                                from decimal import Decimal
                                margen_decimal = Decimal(str(producto.margen_ganancia))
                                precio_calculado = costo_calculado * (Decimal('1') + margen_decimal / Decimal('100'))
                                producto.precio_venta_calculado = precio_calculado

                            # El método save() del modelo se encargará de la sincronización de precios
                            producto.save()

                    # Mensaje de éxito personalizado
                    if nueva_categoria:
                        messages.success(request, f'Producto creado exitosamente con la nueva categoría "{nueva_categoria}".')
                    else:
                        messages.success(request, 'Producto creado exitosamente.')

                    return redirect('gestion:lista_productos')

            except Exception as e:
                messages.error(request, f'Error al crear el producto: {str(e)}')
                return render(request, 'modules/productos/form.html', {'form': form, 'title': 'Crear Producto', 'producto': None})
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, error)
                    else:
                        field_label = form.fields[field].label if field in form.fields else field
                        messages.error(request, f'{field_label}: {error}')
            return render(request, 'modules/productos/form.html', {'form': form, 'title': 'Crear Producto', 'producto': None})
    else:
        form = ProductoForm()

    return render(request, 'modules/productos/form.html', {'form': form, 'title': 'Crear Producto', 'producto': None})


@login_required
def detalle_producto(request, pk):
    """Vista de detalle de producto con información completa."""
    producto = get_object_or_404(Producto, pk=pk)

    # Calcular estadísticas básicas (solo ventas activas)
    ventas_mes = VentaDetalle.objects.filter(
        producto=producto,
        venta__fecha__month=timezone.now().month,
        venta__fecha__year=timezone.now().year,
        venta__eliminada=False  # ✅ Excluir ventas eliminadas
    ).aggregate(
        total_vendido=models.Sum('cantidad'),
        total_ventas=models.Count('venta', distinct=True)
    )

    # Obtener últimas ventas del producto (solo activas)
    ultimas_ventas = VentaDetalle.objects.filter(
        producto=producto,
        venta__eliminada=False  # ✅ Excluir ventas eliminadas
    ).select_related('venta').order_by('-venta__fecha')[:5]

    context = {
        'producto': producto,
        'ventas_mes': ventas_mes['total_ventas'] or 0,
        'cantidad_vendida_mes': ventas_mes['total_vendido'] or 0,
        'ultimas_ventas': ultimas_ventas,
    }

    return render(request, 'modules/productos/detalle.html', context)


@login_required
def editar_producto(request, pk):
    """Vista para editar un producto existente."""
    producto = get_object_or_404(Producto, pk=pk)

    if not request.user.has_perm('gestion.change_producto'):
        messages.error(request, 'No tienes permiso para editar productos.')
        return redirect('gestion:lista_productos')

    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Guardar stock anterior para comparar
                    stock_anterior = producto.stock

                    # Guardar el producto
                    producto = form.save(commit=False)

                    # Solo procesar producción si hay cantidad_a_producir explícita
                    # (campo separado del stock, para producir desde materias primas)
                    cantidad_a_producir = form.cleaned_data.get('cantidad_a_producir', 0)
                    if cantidad_a_producir and cantidad_a_producir > 0:
                        # PRODUCCIÓN: Descontar materias primas y aumentar stock
                        ok, faltantes = producto.verificar_stock_materias_primas(cantidad_a_producir)
                        if not ok:
                            faltantes_str = ", ".join([f"{f['materia_prima']} (necesaria: {f['necesaria']} {f['unidad']}, disponible: {f['disponible']})" for f in faltantes])
                            messages.error(request, f"No hay suficiente stock de materias primas para producir {cantidad_a_producir} unidades: {faltantes_str}")
                            raise Exception("Stock de materias primas insuficiente")
                        producto.descontar_materias_primas(cantidad_a_producir, request.user)
                        producto.stock += cantidad_a_producir
                        messages.info(request, f'✅ Producidas {cantidad_a_producir} unidades desde materias primas')

                    # Si solo se cambió el stock manualmente (sin producir)
                    # simplemente guardar el nuevo valor sin descontar nada
                    producto.save()

                    messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente.')
                return redirect('gestion:lista_productos')
            except Exception as e:
                messages.error(request, f'Error inesperado al actualizar el producto: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
            messages.error(request, 'Error al actualizar el producto. Verifica los datos ingresados.')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'modules/productos/form.html', {
        'form': form,
        'title': 'Editar Producto',
        'producto': producto
    })


@login_required
def eliminar_producto(request, pk):
    """Vista para eliminar un producto."""
    producto = get_object_or_404(Producto, pk=pk)

    # Verificar permisos
    if not request.user.has_perm('gestion.delete_producto'):
        messages.error(request, 'No tienes permiso para eliminar productos.')
        return redirect('gestion:lista_productos')

    if request.method == 'POST':
        try:
            nombre_producto = producto.nombre
            categoria_producto = producto.get_categoria_display()
            stock_producto = producto.stock

            with transaction.atomic():
                # Verificar si el producto tiene ventas asociadas
                ventas_asociadas = producto.ventadetalle_set.count()
                if ventas_asociadas > 0:
                    messages.warning(
                        request,
                        f'El producto "{nombre_producto}" tiene {ventas_asociadas} venta(s) asociada(s). '
                        'Se eliminará el producto pero se mantendrá el historial de ventas.'
                    )

                # Eliminar el producto
                producto.delete()

                # Mensaje de éxito con información detallada
                messages.success(
                    request,
                    f'Producto "{nombre_producto}" (Categoría: {categoria_producto}, Stock: {stock_producto}) '
                    'eliminado exitosamente.'
                )

            return redirect('gestion:lista_productos')

        except Exception as e:
            messages.error(request, f'Error al eliminar el producto: {str(e)}')
            return redirect('gestion:lista_productos')

    # Para GET, mostrar página de confirmación
    context = {
        'producto': producto,
        'ventas_asociadas': producto.ventadetalle_set.count()
    }
    return render(request, 'modules/productos/confirmar_eliminar.html', context)


@login_required
def exportar_productos(request):
    """Exporta productos a Excel."""
    if not request.user.has_perm('gestion.export_producto'):
        messages.error(request, 'No tienes permiso para exportar productos.')
        return redirect('gestion:lista_productos')
    try:
        producto_resource = ProductoResource()
        dataset = producto_resource.export()
        response = HttpResponse(dataset.xlsx, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="productos_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        # Auditoría: registrar acción
        # LogProducto.objects.create(usuario=request.user, accion='exportar', descripción='Exportación de productos a Excel')
        return response
    except Exception as e:
        messages.error(request, f'Error inesperado al exportar productos: {str(e)}')
        return redirect('gestion:lista_productos')


# ==================== VISTAS DE MATERIAS PRIMAS ====================

@login_required
def lista_materias_primas(request):
    """Vista para listar materias primas con filtros."""
    materias_primas = MateriaPrima.objects.filter(activo=True)
    query = request.GET.get('q')
    proveedor = request.GET.get('proveedor')
    estado_stock = request.GET.get('estado_stock')

    if query:
        materias_primas = materias_primas.filter(
            Q(nombre__icontains=query) |
            Q(descripcion__icontains=query) |
            Q(proveedor__icontains=query)
        )

    if proveedor:
        materias_primas = materias_primas.filter(proveedor__icontains=proveedor)

    if estado_stock:
        if estado_stock == 'agotado':
            materias_primas = materias_primas.filter(stock_actual=0)
        elif estado_stock == 'bajo':
            materias_primas = materias_primas.filter(stock_actual__gt=0, stock_actual__lte=F('stock_minimo'))
        elif estado_stock == 'normal':
            materias_primas = materias_primas.filter(stock_actual__gt=F('stock_minimo'))

    # Estadísticas para KPIs
    total_materias = MateriaPrima.objects.filter(activo=True).count()
    con_stock = MateriaPrima.objects.filter(activo=True, stock_actual__gt=0).count()
    stock_bajo = MateriaPrima.objects.filter(
        activo=True,
        stock_actual__gt=0,
        stock_actual__lte=F('stock_minimo')
    ).count()
    valor_total = MateriaPrima.objects.filter(activo=True).aggregate(
        total=Sum(F('stock_actual') * F('costo_unitario'))
    )['total'] or 0

    stats = {
        'con_stock': con_stock,
        'stock_bajo': stock_bajo,
        'valor_total': valor_total
    }

    # Obtener proveedores únicos para el filtro
    proveedores = MateriaPrima.objects.filter(activo=True).values_list('proveedor', flat=True).distinct().exclude(proveedor__isnull=True).exclude(proveedor='')

    context = {
        'materias_primas': materias_primas,
        'proveedores': proveedores,
        'stats': stats,
        'query': query or '',
        'proveedor_seleccionado': proveedor or '',
        'estado_stock_seleccionado': estado_stock or '',
    }

    return render(request, 'gestion/materias_primas/lista_simple.html', context)


@login_required
def lista_inventario(request):
    """Vista de inventario optimizada - usa InventarioService para KPIs inteligentes."""
    try:
        from django.core.paginator import Paginator

        from gestion.services.inventario_service import InventarioService

        # Inicializar servicio de inventario
        service = InventarioService()

        # Obtener KPIs inteligentes del servicio
        kpis = service.get_kpis_inventario()

        # Reutilizar lógica existente con paginación
        materias_primas = MateriaPrima.objects.filter(activo=True).order_by('nombre')
        query = request.GET.get('q', '')
        proveedor_seleccionado = request.GET.get('proveedor', '')
        estado_stock = request.GET.get('estado_stock', '')

        # Aplicar filtros (lógica reutilizada)
        if query:
            materias_primas = materias_primas.filter(
                Q(nombre__icontains=query) |
                Q(descripcion__icontains=query) |
                Q(proveedor__icontains=query)
            )

        if proveedor_seleccionado:
            materias_primas = materias_primas.filter(proveedor__icontains=proveedor_seleccionado)

        if estado_stock:
            if estado_stock == 'agotado':
                materias_primas = materias_primas.filter(stock_actual=0)
            elif estado_stock == 'bajo':
                materias_primas = materias_primas.filter(
                    stock_actual__gt=0,
                    stock_actual__lte=F('stock_minimo')
                )
            elif estado_stock == 'normal':
                materias_primas = materias_primas.filter(stock_actual__gt=F('stock_minimo'))

        # Proveedores únicos para filtros
        all_materias = MateriaPrima.objects.filter(activo=True)
        proveedores = (all_materias.values_list('proveedor', flat=True)
                      .distinct()
                      .exclude(proveedor__isnull=True)
                      .exclude(proveedor='')
                      .order_by('proveedor'))

        total_proveedores = len(set(proveedores))

        # Paginación eficiente
        paginator = Paginator(materias_primas, 25)
        page_number = request.GET.get('page', 1)
        materias_paginadas = paginator.get_page(page_number)

        context = {
            'materias_primas': materias_paginadas,
            'proveedores': proveedores,
            'total_proveedores': total_proveedores,
            # KPIs inteligentes del servicio
            'kpis': kpis,
            # KPIs legacy para compatibilidad
            'total_materias': all_materias.count(),
            'con_stock': all_materias.filter(stock_actual__gt=0).count(),
            'stock_bajo': kpis['stock_critico']['cantidad'],
            'stock_critico': kpis['stock_critico']['cantidad'],
            'valor_total': kpis['valor_total']['valor'],
        }

        return render(request, 'modules/inventario/lista_inventario.html', context)

    except Exception as e:
        messages.error(request, f'Error al cargar inventario: {str(e)}')
        return redirect('gestion:panel_control')


@login_required
def crear_materia_prima(request):
    """Vista para crear nueva materia prima."""
    if not request.user.has_perm('gestion.add_materiaprima'):
        messages.error(request, 'No tienes permiso para crear materias primas.')
        return redirect('gestion:lista_inventario')
    if request.method == 'POST':
        form = MateriaPrimaForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    materia_prima = form.save()
                    # Registrar movimiento inicial si hay stock
                    if materia_prima.stock_actual > 0:
                        MovimientoMateriaPrima.objects.create(
                            materia_prima=materia_prima,
                            tipo_movimiento='entrada',
                            cantidad=materia_prima.stock_actual,
                            cantidad_anterior=0,
                            cantidad_nueva=materia_prima.stock_actual,
                            motivo='Stock inicial',
                            usuario=request.user
                        )
                    # Auditoría: registrar acción
                    # LogMateriaPrima.objects.create(usuario=request.user, accion='crear', materia_prima=materia_prima)
                messages.success(request, f'Materia prima "{materia_prima.nombre}" creada exitosamente.')
                return redirect('gestion:lista_inventario')
            except Exception as e:
                messages.error(request, f'Error inesperado al crear la materia prima: {str(e)}')
        else:
            messages.error(request, 'Error al crear la materia prima. Verifica los datos.')
    else:
        form = MateriaPrimaForm()
    return render(request, 'modules/materias_primas/materias_primas/crear.html', {
        'form': form,
        'titulo': 'Crear Materia Prima'
    })


@login_required
def editar_materia_prima(request, pk):
    """Vista para editar materia prima."""
    materia_prima = get_object_or_404(MateriaPrima, pk=pk)
    if not request.user.has_perm('gestion.change_materiaprima'):
        messages.error(request, 'No tienes permiso para editar materias primas.')
        return redirect('gestion:lista_inventario')
    stock_anterior = materia_prima.stock_actual
    if request.method == 'POST':
        form = MateriaPrimaForm(request.POST, instance=materia_prima)
        if form.is_valid():
            try:
                with transaction.atomic():
                    materia_prima = form.save()
                    # Registrar movimiento si cambió el stock
                    if stock_anterior != materia_prima.stock_actual:
                        diferencia = materia_prima.stock_actual - stock_anterior
                        tipo_mov = 'entrada' if diferencia > 0 else 'salida'
                        MovimientoMateriaPrima.objects.create(
                            materia_prima=materia_prima,
                            tipo_movimiento='ajuste',
                            cantidad=abs(diferencia),
                            cantidad_anterior=stock_anterior,
                            cantidad_nueva=materia_prima.stock_actual,
                            motivo=f'Ajuste manual - {tipo_mov}',
                            usuario=request.user
                        )
                    # Auditoría: registrar acción
                    # LogMateriaPrima.objects.create(usuario=request.user, accion='editar', materia_prima=materia_prima)
                messages.success(request, f'Materia prima "{materia_prima.nombre}" actualizada exitosamente.')
                return redirect('gestion:lista_inventario')
            except Exception as e:
                messages.error(request, f'Error inesperado al actualizar la materia prima: {str(e)}')
        else:
            messages.error(request, 'Error al actualizar la materia prima. Verifica los datos.')
    else:
        form = MateriaPrimaForm(instance=materia_prima)
    return render(request, 'modules/materias_primas/materias_primas/form.html', {
        'form': form,
        'titulo': 'Editar Materia Prima',
        'materia_prima': materia_prima
    })


@login_required
def detalle_materia_prima(request, pk):
    """Vista detalle de materia prima con movimientos, lotes FIFO, permisos y errores."""
    if not request.user.has_perm('gestion.view_materiaprima'):
        messages.error(request, 'No tienes permiso para ver detalles de materias primas.')
        return redirect('gestion:lista_inventario')
    try:
        materia_prima = get_object_or_404(MateriaPrima, pk=pk)
        movimientos = materia_prima.movimientos.all()[:20]  # Últimos 20 movimientos
        productos_relacionados = ProductoMateriaPrima.objects.filter(materia_prima=materia_prima)
        # Importar el modelo de lotes y obtener los lotes FIFO de esta materia prima
        lotes = LoteMateriaPrima.objects.filter(materia_prima=materia_prima).order_by('fecha_entrada', 'id')
        context = {
            'materia_prima': materia_prima,
            'movimientos': movimientos,
            'productos_relacionados': productos_relacionados,
            'lotes': lotes,
        }
        # Auditoría: registrar acción
        # LogMateriaPrima.objects.create(usuario=request.user, accion='ver', materia_prima=materia_prima, descripcion='Detalle de materia prima')
        return render(request, 'modules/materias_primas/materias_primas/detalle.html', context)
    except Exception as e:
        messages.error(request, f'Error inesperado al mostrar detalle: {str(e)}')
        return redirect('gestion:lista_inventario')


@login_required
def movimiento_materia_prima(request, pk):
    """Vista para registrar movimiento de materia prima con permisos, errores y auditoría."""
    if not request.user.has_perm('gestion.change_materiaprima'):
        messages.error(request, 'No tienes permiso para registrar movimientos de materias primas.')
        return redirect('gestion:lista_inventario')
    try:
        materia_prima = get_object_or_404(MateriaPrima, pk=pk)
        if request.method == 'POST':
            form = MovimientoMateriaPrimaForm(request.POST)
            if form.is_valid():
                movimiento = form.save(commit=False)
                movimiento.usuario = request.user
                movimiento.cantidad_anterior = materia_prima.stock_actual
                # Actualizar stock según tipo de movimiento
                if movimiento.tipo_movimiento in ['entrada', 'devolucion']:
                    materia_prima.stock_actual += movimiento.cantidad
                elif movimiento.tipo_movimiento in ['salida', 'produccion']:
                    if materia_prima.stock_actual >= movimiento.cantidad:
                        # FIFO: descontar de lotes
                        cantidad_restante = float(movimiento.cantidad)
                        lotes = LoteMateriaPrima.objects.filter(materia_prima=materia_prima, cantidad_disponible__gt=0).order_by('fecha_entrada', 'id')
                        for lote in lotes:
                            if cantidad_restante <= 0:
                                break
                            disponible = float(lote.cantidad_disponible)
                            a_consumir = min(disponible, cantidad_restante)
                            lote.cantidad_disponible = disponible - a_consumir
                            if lote.cantidad_disponible == 0:
                                lote.fecha_consumo = timezone.now().date()
                            lote.save()
                            cantidad_restante -= a_consumir
                        materia_prima.stock_actual -= movimiento.cantidad
                    else:
                        messages.error(request, 'No hay suficiente stock disponible.')
                        return render(request, 'modules/materias_primas/materias_primas/movimiento.html', {
                            'form': form,
                            'materia_prima': materia_prima
                        })
                elif movimiento.tipo_movimiento == 'ajuste':
                    materia_prima.stock_actual = movimiento.cantidad
                movimiento.cantidad_nueva = materia_prima.stock_actual
                # Guardar ambos objetos
                with transaction.atomic():
                    materia_prima.save()
                    movimiento.save()
                # Auditoría: registrar acción
                # LogMateriaPrima.objects.create(usuario=request.user, accion='movimiento', materia_prima=materia_prima, descripcion=f'Movimiento: {movimiento.tipo_movimiento}')
                messages.success(request, f'Movimiento registrado exitosamente. Nuevo stock: {materia_prima.stock_actual}')
                return redirect('gestion:detalle_materia_prima', pk=pk)
            else:
                messages.error(request, 'Error al registrar el movimiento. Verifica los datos.')
        else:
            form = MovimientoMateriaPrimaForm(initial={'materia_prima': materia_prima})
        return render(request, 'modules/materias_primas/materias_primas/movimiento.html', {
            'form': form,
            'materia_prima': materia_prima
        })
    except Exception as e:
        messages.error(request, f'Error inesperado al registrar movimiento: {str(e)}')
        return redirect('gestion:detalle_materia_prima', pk=pk)


@login_required
def exportar_materias_primas_excel(request):
    """Exporta materias primas a Excel con formato y estilos."""
    if not request.user.has_perm('gestion.export_materiaprima'):
        messages.error(request, 'No tienes permiso para exportar materias primas.')
        return redirect('gestion:lista_inventario')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Materias Primas"

        # Headers
        headers = ['ID', 'Nombre', 'Descripción', 'Proveedor', 'Costo Unitario', 'Stock Actual', 'Stock Mínimo', 'Unidad Medida', 'Valor Total']
        ws.append(headers)

        # Dar formato a headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Agregar datos
        materias_primas = MateriaPrima.objects.filter(activo=True)
        for mp in materias_primas:
            ws.append([
                mp.id,
                mp.nombre,
                mp.descripcion or '',
                mp.proveedor or '',
                float(mp.costo_unitario),
                float(mp.stock_actual),
                float(mp.stock_minimo),
                mp.unidad_medida,
                float(mp.valor_total_stock),
            ])

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15

        # Crear respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="materias_primas_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)

        return response
    except Exception as e:
        messages.error(request, f'Error al exportar materias primas: {str(e)}')
        return redirect('gestion:lista_inventario')


# ==================== REPORTES DE PRODUCTOS E INGREDIENTES ====================

@login_required
def reporte_stock_materias_primas(request):
    """Reporte detallado de stock de materias primas."""
    if not request.user.has_perm('gestion.view_reporte'):
        messages.error(request, 'No tienes permiso para ver reportes de stock de materias primas.')
        return redirect('gestion:panel_control')
    try:
        materias_primas = MateriaPrima.objects.filter(activo=True)
        # Clasificación de materias primas según estado de stock
        stock_critico = materias_primas.filter(stock_actual__lte=models.F('stock_minimo'))
        stock_bajo = materias_primas.filter(
            stock_actual__gt=models.F('stock_minimo'),
            stock_actual__lte=models.F('stock_minimo') * 2
        )
        stock_normal = materias_primas.exclude(
            id__in=list(stock_critico.values_list('id', flat=True)) +
                   list(stock_bajo.values_list('id', flat=True))
        )
        # Cálculo de valores totales y críticos
        valor_total = sum([mp.valor_total_stock for mp in materias_primas])
        valor_critico = sum([mp.valor_total_stock for mp in stock_critico])
        context = {
            'materias_primas': materias_primas,
            'stock_critico': stock_critico,
            'stock_bajo': stock_bajo,
            'stock_normal': stock_normal,
            'valor_total': valor_total,
            'valor_critico': valor_critico,
            'total_materias': materias_primas.count(),
        }
        # Auditoría: acción de consulta de reporte de stock (descomentar si se implementa logging)
        # LogReporte.objects.create(usuario=request.user, accion='ver', descripción='Reporte de stock de materias primas')
        return render(request, 'gestion/reportes/stock_materias_primas.html', context)
    except Exception as e:
        messages.error(request, f'Error inesperado al generar el reporte de stock: {str(e)}')
        return redirect('gestion:panel_control')


@login_required
def reporte_costos_produccion(request):
    """Reporte de costos de producción por producto."""
    if not request.user.has_perm('gestion.view_reporte'):
        messages.error(request, 'No tienes permiso para ver reportes de costos de producción.')
        return redirect('gestion:panel_control')
    try:
        # Analiza productos con receta y calcula margen de ganancia
        productos_con_receta = Producto.objects.filter(recetas__isnull=False).distinct()
        productos_analisis = []
        for producto in productos_con_receta:
            costo_materias = producto.costo_materias_primas
            margen_ganancia = producto.precio - costo_materias
            porcentaje_margen = (margen_ganancia / producto.precio * 100) if producto.precio > 0 else 0
            productos_analisis.append({
                'producto': producto,
                'costo_materias': costo_materias,
                'precio_venta': producto.precio,
                'margen_ganancia': margen_ganancia,
                'porcentaje_margen': porcentaje_margen,
            })
        productos_analisis.sort(key=lambda x: x['porcentaje_margen'], reverse=True)
        context = {
            'productos_analisis': productos_analisis,
            'total_productos': len(productos_analisis),
        }
        return render(request, 'gestion/reportes/costos_produccion.html', context)
    except Exception as e:
        messages.error(request, f'Error inesperado al generar el reporte de costos: {str(e)}')
        return redirect('gestion:panel_control')

# ==================== IMPORTS PRINCIPALES ====================
import json
import logging
import traceback
from datetime import datetime, timedelta
from decimal import Decimal

logger = logging.getLogger(__name__)

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import models, transaction
from django.db.models import Count, F, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django_ratelimit.decorators import ratelimit

from .analytics import AnalyticsRentabilidad
from .forms import (
    AjusteMateriaPrimaForm,
    AjusteProductoForm,
    CompraForm,
    MateriaPrimaForm,
    MovimientoMateriaPrimaForm,
    ProductoForm,
    RecetaForm,
    VentaConMateriasForm,
    VentaDetalleFormSet,
    VentaForm,
)

# 🔧 CORREGIDO: Eliminados imports duplicados (Decimal y timezone)
# y mantuvimos models porque se usa para models.Sum, models.Count, models.F
# ==================== IMPORTS PARA LOGGING ROBUSTO ====================
from .logging_system import LinoLogger, get_request_info, log_business_operation
from .models import (
    AjusteInventario,
    Compra,
    MateriaPrima,
    MovimientoMateriaPrima,
    Producto,
    ProductoMateriaPrima,
    Receta,
    RecetaMateriaPrima,
    Venta,
    VentaDetalle,
)
from .resources import ProductoResource, VentaResource


@login_required
def editar_receta(request, pk):
    """Vista para editar una receta existente."""
    receta = get_object_or_404(Receta, pk=pk)

    if not request.user.has_perm('gestion.change_receta'):
        messages.error(request, 'No tienes permiso para editar recetas.')
        return redirect('gestion:lista_recetas')

    if request.method == 'POST':
        form = RecetaForm(request.POST, instance=receta)
        if form.is_valid():
            try:
                with transaction.atomic():
                    receta = form.save(commit=False)
                    receta.save()

                    # Eliminar ingredientes existentes
                    RecetaMateriaPrima.objects.filter(receta=receta).delete()

                    # Procesar nuevos ingredientes
                    procesar_ingredientes_receta(request.POST, receta)

                    messages.success(request, 'Receta actualizada exitosamente.')
                    return redirect('gestion:lista_recetas')
            except Exception as e:
                messages.error(request, f'Error al actualizar la receta: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{form.fields[field].label}: {error}')
    else:
        form = RecetaForm(instance=receta)

    # Obtener ingredientes actuales para mostrar en el formulario
    ingredientes_actuales = []
    for ingrediente in receta.recetamateriaprima_set.all():
        ingredientes_actuales.append({
            'materia_prima_id': ingrediente.materia_prima.id,
            'materia_prima_nombre': ingrediente.materia_prima.nombre,
            'cantidad': float(ingrediente.cantidad),
            'unidad': ingrediente.unidad,
            'costo': float(ingrediente.costo_ingrediente())
        })

    context = {
        'form': form,
        'receta': receta,
        'materias_primas': MateriaPrima.objects.filter(activo=True),
        'ingredientes_actuales': json.dumps(ingredientes_actuales),
        'es_edicion': True
    }
    return render(request, 'modules/recetas/form.html', context)

@login_required
def eliminar_receta(request, pk):
    """Vista para eliminar una receta."""
    receta = get_object_or_404(Receta, pk=pk)

    if not request.user.has_perm('gestion.delete_receta'):
        messages.error(request, 'No tienes permiso para eliminar recetas.')
        return redirect('gestion:lista_recetas')

    if request.method == 'POST':
        try:
            # Verificar si la receta está siendo usada por productos
            productos_usando = receta.productos.all()
            if productos_usando.exists():
                productos_nombres = ', '.join([p.nombre for p in productos_usando[:3]])
                if productos_usando.count() > 3:
                    productos_nombres += f" y {productos_usando.count() - 3} más"
                messages.error(
                    request,
                    f'No se puede eliminar la receta porque está siendo usada por los productos: {productos_nombres}'
                )
                return redirect('gestion:lista_recetas')

            nombre_receta = receta.nombre
            receta.delete()
            messages.success(request, f'Receta "{nombre_receta}" eliminada exitosamente.')

        except Exception as e:
            messages.error(request, f'Error al eliminar la receta: {str(e)}')

        return redirect('gestion:lista_recetas')

    # Para GET, mostrar página de confirmación
    context = {
        'receta': receta,
        'productos_usando': receta.productos.all(),
        'ingredientes_count': receta.recetamateriaprima_set.count()
    }
    return render(request, 'modules/recetas/confirmar_eliminar_receta.html', context)

@login_required
def detalle_receta(request, pk):
    """Vista para ver el detalle de una receta."""
    receta = get_object_or_404(Receta, pk=pk)

    # Calcular información adicional
    total_ingredientes = receta.recetamateriaprima_set.count()
    costo_total = receta.costo_total()
    productos_usando = receta.productos.all()

    # Obtener ingredientes con información detallada
    ingredientes = []
    for ingrediente in receta.recetamateriaprima_set.all():
        ingredientes.append({
            'materia_prima': ingrediente.materia_prima,
            'cantidad': ingrediente.cantidad,
            'unidad': ingrediente.unidad,
            'costo_unitario': ingrediente.materia_prima.costo_unitario,
            'costo_total': ingrediente.costo_ingrediente(),
            'porcentaje_costo': (ingrediente.costo_ingrediente() / costo_total * 100) if costo_total > 0 else 0
        })

    # Preparar subtitle con estado e ingredientes
    estado_text = "Receta Activa" if receta.activa else "Receta Inactiva"
    subtitle_text = f"{estado_text} • {total_ingredientes} ingrediente(s)"

    context = {
        'receta': receta,
        'ingredientes': ingredientes,
        'total_ingredientes': total_ingredientes,
        'costo_total': costo_total,
        'productos_usando': productos_usando,
        'puede_editar': request.user.has_perm('gestion.change_receta'),
        'puede_eliminar': request.user.has_perm('gestion.delete_receta'),
        'subtitle_text': subtitle_text,
    }
    return render(request, 'modules/recetas/detalle.html', context)

# ==================== API: PRECIO DE PRODUCTO ====================
@login_required
def producto_precio(request, pk):
    """Devuelve el precio de un producto en formato JSON. Para uso en API/ajax."""
    try:
        producto = Producto.objects.get(pk=pk)
        return JsonResponse({'success': True, 'precio': float(producto.precio)})
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ==================== VISTAS BASE PARA SECCIONES LATERALES ====================
@login_required
def gastos_inversiones(request):
    return render(request, 'gestion/gastos_inversiones.html')

@login_required
def usuarios(request):
    context = {
        'title': 'Usuarios',
        'subtitle': 'Gestión de usuarios y permisos del sistema',
        'icon': 'people',
    }
    return render(request, 'modules/configuracion/usuarios.html', context)

@login_required
def configuracion(request):
    context = {
        'title': 'Configuración',
        'subtitle': 'Configuración general del sistema',
        'icon': 'gear',
    }
    return render(request, 'modules/configuracion/panel.html', context)

# ==================== VISTAS DE COMPRAS AL MAYOREO ====================
@login_required
def panel_control_original(request):
    try:
        # Estadísticas de productos
        total_productos = Producto.objects.count()
        productos_stock_bajo = Producto.objects.filter(stock__lte=F('stock_minimo'))
        productos_stock_bajo_count = productos_stock_bajo.count()

        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)
        ventas_mes = Venta.objects.filter(fecha__date__gte=inicio_mes).count()
        ingresos_mes = Venta.objects.filter(fecha__date__gte=inicio_mes).aggregate(
            total=Sum('total')
        )['total'] or 0

        # NUEVAS ESTADÍSTICAS DE COMPRAS
        total_compras = Compra.objects.count()
        compras_mes = Compra.objects.filter(fecha_compra__gte=inicio_mes).count()
        inversion_mes = Compra.objects.filter(fecha_compra__gte=inicio_mes).aggregate(
            total=Sum('precio_mayoreo')
        )['total'] or 0
        inversion_total = Compra.objects.aggregate(
            total=Sum('precio_mayoreo')
        )['total'] or 0

        # Stock bajo de materias primas (usando MateriaPrima)
        materias_stock_bajo = MateriaPrima.objects.filter(stock_actual__lte=F('stock_minimo')).count()
        total_materias_primas = MateriaPrima.objects.count()
        materias_stock_critico = MateriaPrima.objects.filter(stock_actual__lte=F('stock_minimo')).count()
        valor_inventario_materias = MateriaPrima.objects.aggregate(
            total=Sum(F('stock_actual') * F('costo_unitario'))
        )['total'] or 0
        materias_primas_criticas = MateriaPrima.objects.filter(stock_actual__lte=F('stock_minimo')).order_by('stock_actual')[:10]

        # Últimas actividades
        ultimas_ventas = Venta.objects.prefetch_related('detalles__producto').order_by('-fecha')[:5]
        ultimas_compras = Compra.objects.order_by('-fecha_compra')[:5]

        # Datos para gráfico de ventas de los últimos 7 días
        from datetime import timedelta
        ventas_ultimos_7_dias = []
        labels_7_dias = []
        for i in range(6, -1, -1):
            fecha = hoy - timedelta(days=i)
            ventas_dia = Venta.objects.filter(fecha__date=fecha).aggregate(
                total=Sum('total')
            )['total'] or 0
            ventas_ultimos_7_dias.append(float(ventas_dia))
            labels_7_dias.append(fecha.strftime('%d/%m'))

        # Comparación con mes anterior
        mes_anterior = inicio_mes - timedelta(days=30)
        fin_mes_anterior = inicio_mes - timedelta(days=1)
        ventas_mes_anterior = Venta.objects.filter(
            fecha__date__gte=mes_anterior,
            fecha__date__lte=fin_mes_anterior
        ).count()
        ingresos_mes_anterior = Venta.objects.filter(
            fecha__date__gte=mes_anterior,
            fecha__date__lte=fin_mes_anterior
        ).aggregate(total=Sum('total'))['total'] or 0

        # Calcular tendencias de manera más inteligente
        if ventas_mes_anterior > 0:
            tendencia_ventas = ((ventas_mes - ventas_mes_anterior) / ventas_mes_anterior) * 100
        elif ventas_mes > 0:
            tendencia_ventas = 100  # 100% de aumento desde 0
        else:
            tendencia_ventas = 0  # Sin cambio (ambos son 0)

        if ingresos_mes_anterior > 0:
            tendencia_ingresos = ((ingresos_mes - ingresos_mes_anterior) / ingresos_mes_anterior) * 100
        elif ingresos_mes > 0:
            tendencia_ingresos = 100  # 100% de aumento desde 0
        else:
            tendencia_ingresos = 0  # Sin cambio (ambos son 0)

        # Limitar las tendencias a valores razonables (máximo ±999%)
        tendencia_ventas = max(-999, min(999, tendencia_ventas))
        tendencia_ingresos = max(-999, min(999, tendencia_ingresos))

        # Productos más vendidos
        productos_mas_vendidos = VentaDetalle.objects.filter(
            venta__fecha__date__gte=inicio_mes
        ).values('producto__nombre').annotate(
            total_vendido=Sum('cantidad')
        ).order_by('-total_vendido')[:5]

        # Margen de ganancia estimado (simplificado)
        margen_bruto = ingresos_mes - inversion_mes
        porcentaje_margen = (margen_bruto / max(ingresos_mes, 1)) * 100 if ingresos_mes > 0 else 0

        # Alertas críticas
        productos_sin_stock = Producto.objects.filter(stock=0).count()
        materias_vencen_pronto = MateriaPrima.objects.filter(
            stock_actual__gt=0,
            stock_actual__lte=F('stock_minimo') * 2
        ).count()

        # ==================== INTEGRACIÓN DE RENTABILIDAD ====================
        try:
            from .analytics import AnalyticsRentabilidad
            analytics = AnalyticsRentabilidad()

            # KPIs de rentabilidad para el dashboard principal
            kpis_rentabilidad = analytics.get_kpis_rentabilidad()
            alertas_rentabilidad = analytics.get_alertas_rentabilidad()

            # Productos en pérdida y críticos para mostrar en dashboard
            productos_perdida = kpis_rentabilidad['productos_en_perdida']
            productos_criticos_rentabilidad = kpis_rentabilidad['productos_criticos']
            margen_promedio_negocio = kpis_rentabilidad['margen_promedio_ponderado']

            # Top 3 alertas más críticas para mostrar en dashboard
            alertas_criticas_dashboard = [a for a in alertas_rentabilidad if a['severidad'] == 'critica'][:3]

        except Exception:
            # Si hay error en analytics, continuar sin rentabilidad
            productos_perdida = 0
            productos_criticos_rentabilidad = 0
            margen_promedio_negocio = 0
            alertas_criticas_dashboard = []

        stock_normal = max(total_productos - productos_stock_bajo_count, 0)
        context = {
            # Productos
            'total_productos': total_productos,
            'productos_stock_bajo': productos_stock_bajo[:10],
            'productos_stock_bajo_count': productos_stock_bajo_count,
            'productos_sin_stock': productos_sin_stock,
            # Ventas
            'ventas_mes': ventas_mes,
            'ingresos_mes': ingresos_mes,
            'ultimas_ventas': ultimas_ventas,
            'tendencia_ventas': round(tendencia_ventas, 1),
            'tendencia_ingresos': round(tendencia_ingresos, 1),
            'productos_mas_vendidos': productos_mas_vendidos,
            # Finanzas
            'margen_bruto': margen_bruto,
            'porcentaje_margen': round(porcentaje_margen, 1),
            # Datos para gráfico
            'ventas_labels': json.dumps(labels_7_dias),
            'ventas_data': json.dumps(ventas_ultimos_7_dias),
            # Compras
            'total_compras': total_compras,
            'compras_mes': compras_mes,
            'inversion_mes': inversion_mes,
            'inversion_total': inversion_total,
            'materias_stock_bajo': materias_stock_bajo,
            'ultimas_compras': ultimas_compras,
            # Materias primas
            'total_materias_primas': total_materias_primas,
            'materias_stock_critico': materias_stock_critico,
            'materias_vencen_pronto': materias_vencen_pronto,
            'valor_inventario_materias': valor_inventario_materias,
            'materias_primas_criticas': materias_primas_criticas,
            'stock_normal': stock_normal,
            # Rentabilidad integrada
            'productos_perdida': productos_perdida,
            'productos_criticos_rentabilidad': productos_criticos_rentabilidad,
            'margen_promedio_negocio': margen_promedio_negocio,
            'alertas_criticas_dashboard': alertas_criticas_dashboard,
        }
        return render(request, 'gestion/dashboard.html', context)
    except Exception as e:
        return render(request, 'gestion/error_panel_control.html', {
            'error_message': f'Error inesperado al cargar el panel de control: {str(e)}'
        })

@login_required
def panel_control_clean(request):
    """Vista del dashboard limpio sin JavaScript problemático."""
    try:
        # Estadísticas básicas
        total_productos = Producto.objects.count()
        productos_bajo_stock = Producto.objects.filter(stock__lte=F('stock_minimo'))
        productos_bajo_stock_count = productos_bajo_stock.count()
        productos_sin_stock = Producto.objects.filter(stock=0)

        # Fecha actual y mes
        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)

        # Ventas del mes
        ventas_mes = Venta.objects.filter(fecha__date__gte=inicio_mes)
        resumen_ventas_mes = ventas_mes.aggregate(total=Sum('total'), count=Count('id'))
        ventas_recientes = Venta.objects.prefetch_related('detalles__producto').order_by('-fecha')[:5]

        # Valor del inventario simplificado
        valor_inventario = MateriaPrima.objects.aggregate(
            total=Sum(F('stock_actual') * F('costo_unitario'))
        )['total'] or 0

        # Alertas de stock
        alertas_stock = productos_bajo_stock_count + productos_sin_stock.count()

        context = {
            # KPIs principales
            'total_productos': total_productos,
            'productos_bajo_stock': productos_bajo_stock_count,
            'productos_bajo_stock_lista': productos_bajo_stock[:5],  # Solo 5 para mostrar
            'productos_sin_stock': productos_sin_stock,
            'alertas_stock': alertas_stock,
            'valor_inventario': valor_inventario,
            'resumen_ventas_mes': resumen_ventas_mes,
            'ventas_recientes': ventas_recientes,
        }
        return render(request, 'gestion/dashboard_clean.html', context)
    except Exception as e:
        messages.error(request, f'Error al cargar dashboard: {str(e)}')
        return render(request, 'gestion/dashboard_clean.html', {})

@login_required
def panel_control_minimal(request):
    """Vista del dashboard minimalista - solo Bootstrap y CSS V3."""
    try:
        # Estadísticas básicas
        total_productos = Producto.objects.count()
        productos_bajo_stock = Producto.objects.filter(stock__lte=F('stock_minimo'))
        productos_bajo_stock_count = productos_bajo_stock.count()
        productos_sin_stock = Producto.objects.filter(stock=0)

        # Fecha actual y mes
        hoy = timezone.now().date()
        inicio_mes = hoy.replace(day=1)

        # Ventas del mes
        ventas_mes = Venta.objects.filter(fecha__date__gte=inicio_mes)
        resumen_ventas_mes = ventas_mes.aggregate(total=Sum('total'), count=Count('id'))
        ventas_recientes = Venta.objects.prefetch_related('detalles__producto').order_by('-fecha')[:5]

        # Valor del inventario
        valor_inventario = MateriaPrima.objects.aggregate(
            total=Sum(F('stock_actual') * F('costo_unitario'))
        )['total'] or 0

        # Alertas de stock
        alertas_stock = productos_bajo_stock_count + productos_sin_stock.count()

        context = {
            'total_productos': total_productos,
            'productos_bajo_stock': productos_bajo_stock_count,
            'productos_bajo_stock_lista': productos_bajo_stock[:5],
            'productos_sin_stock': productos_sin_stock,
            'alertas_stock': alertas_stock,
            'valor_inventario': valor_inventario,
            'resumen_ventas_mes': resumen_ventas_mes,
            'ventas_recientes': ventas_recientes,
        }
        return render(request, 'gestion/dashboard_minimal.html', context)
    except Exception as e:
        messages.error(request, f'Error al cargar dashboard minimal: {str(e)}')
        return render(request, 'gestion/dashboard_minimal.html', {})

@login_required
def dashboard_inteligente(request):
    """🧠 Dashboard con Inteligencia de Negocio - LINO V3 - Service Layer Architecture."""
    try:
        from gestion.services import AlertasService, DashboardService, MarketingService

        # � INICIALIZAR SERVICIOS
        dashboard_service = DashboardService()
        alertas_service = AlertasService()
        marketing_service = MarketingService()

        # 📊 OBTENER DATOS DEL DASHBOARD (1 llamada centralizada)
        dashboard_data = dashboard_service.get_dashboard_completo()

        # 🔔 OBTENER CONTADOR DE ALERTAS (sin generar nuevas)
        # NOTA: Las alertas se generan manualmente via management command o panel admin
        # No se generan automáticamente para evitar duplicados en cada carga de página
        if request.user.is_authenticated:
            from gestion.models import Alerta
            alertas_no_leidas = Alerta.objects.filter(
                usuario=request.user,
                leida=False,
                archivada=False
            ).count()
        else:
            alertas_no_leidas = 0

        # 📈 MARKETING INTELLIGENCE
        productos_trending = marketing_service.get_productos_trending(limit=3)
        productos_hero = marketing_service.get_hero_products(limit=3)

        # 📊 DATOS PARA GRÁFICOS AVANZADOS
        # Obtener período desde request (por defecto 7 días)
        periodo_dias = int(request.GET.get('periodo', 7))
        comparar_periodo = request.GET.get('comparar', 'false') == 'true'

        ventas_grafico = dashboard_service.get_ventas_por_periodo(
            dias=periodo_dias,
            comparar=comparar_periodo
        )

        top_productos_grafico = dashboard_service.get_top_productos_grafico(
            dias=30,
            limit=5
        )

        # Serializar datos de gráficos a JSON para JavaScript
        import json
        ventas_grafico_json = {
            'labels': ventas_grafico['labels'],
            'datos': ventas_grafico['datos'],
            'total': float(ventas_grafico['total']),
            'promedio': float(ventas_grafico['promedio'])
        }
        if 'datos_anterior' in ventas_grafico:
            ventas_grafico_json['datos_anterior'] = ventas_grafico['datos_anterior']
            ventas_grafico_json['variacion'] = float(ventas_grafico['variacion'])

        top_productos_json = {
            'labels': top_productos_grafico['labels'],
            'ingresos': top_productos_grafico['ingresos']
        }

        # Preparar datos para gráficos (convertir listas a strings CSV)
        kpis = dashboard_data['kpis']
        ventas_sparkline = ','.join(map(str, kpis['ventas_mes']['sparkline']))

        # Sparklines para nuevos KPIs (compras y ganancia no tienen sparkline aún)
        compras_sparkline = ','.join(map(str, kpis.get('compras_mes', {}).get('sparkline', [0]*7)))

        # DEPRECATED: productos e inventario - mantener temporalmente para compatibilidad
        productos_sparkline = ','.join(map(str, kpis.get('productos', {}).get('sparkline', [0]*7)))
        inventario_sparkline = ','.join(map(str, kpis.get('inventario', {}).get('sparkline', [0]*7)))

        # 🎯 CONTEXTO OPTIMIZADO - Todo desde servicios, cero mock data
        context = {
            # KPIs principales con datos reales
            'kpis': dashboard_data['kpis'],
            'resumen_hoy': dashboard_data['resumen_hoy'],
            'actividad_reciente': dashboard_data['actividad_reciente'],
            'top_productos': dashboard_data['top_productos'],

            # Alertas
            'alertas_criticas': alertas_no_leidas,

            # Marketing intelligence
            'productos_trending': productos_trending,
            'productos_hero': productos_hero,

            # Gráficos avanzados (originales para métricas)
            'ventas_grafico': ventas_grafico,
            'top_productos_grafico': top_productos_grafico,

            # Gráficos en JSON para JavaScript
            'ventas_grafico_json': json.dumps(ventas_grafico_json),
            'top_productos_json': json.dumps(top_productos_json),

            'periodo_actual': periodo_dias,
            'comparar_activo': comparar_periodo,

            # Datos para sparklines (formato CSV para Chart.js)
            'ventas_sparkline': ventas_sparkline,
            'compras_sparkline': compras_sparkline,

            # Compatibilidad con template existente
            'ventas_semana': ventas_sparkline,  # Alias
            'total_productos': kpis.get('productos', {}).get('total', 0),  # DEPRECATED - mantener para compatibilidad
            'total_ventas_mes': kpis.get('ventas_mes', {}).get('total', 0),
        }

        return render(request, 'gestion/dashboard_inteligente.html', context)

    except Exception as e:
        messages.error(request, f'Error al cargar dashboard inteligente: {str(e)}')
        logger.error("Error en dashboard_inteligente: %s", e, exc_info=True)

        # Contexto de emergencia con valores seguros
        return render(request, 'gestion/dashboard_inteligente.html', {
            'error': True,
            'kpis': {
                'ventas_mes': {'total': 0, 'variacion': 0, 'sparkline': [0]*7},
                'productos': {'total': 0, 'variacion': 0, 'sparkline': [0]*7},
                'inventario': {'total': 0, 'variacion': 0, 'sparkline': [0]*7},
                'alertas': {'total': 0, 'variacion': 0, 'sparkline': [0]*7},
            },
            'resumen_hoy': {'total_ventas': 0, 'cantidad_ventas': 0, 'productos_vendidos': 0, 'variacion': 0},
            'actividad_reciente': [],
            'top_productos': [],
            'alertas_criticas': 0,
            'productos_trending': [],
            'productos_hero': [],
            'ventas_sparkline': '0,0,0,0,0,0,0',
            'productos_sparkline': '0,0,0,0,0,0,0',
            'inventario_sparkline': '0,0,0,0,0,0,0',
        })

@login_required
def verificar_alertas_stock(request):
    """Función para verificar y mostrar alertas de stock usando stock_minimo personalizado"""
    try:
        productos_agotados = Producto.objects.filter(stock=0)
        productos_criticos = Producto.objects.filter(stock__gt=0, stock__lte=F('stock_minimo'))
        productos_bajos = Producto.objects.filter(stock__gt=F('stock_minimo'), stock__lte=F('stock_minimo') * 2)

        alertas = []
        if productos_agotados.exists():
            alertas.append({
                'tipo': 'danger',
                'titulo': 'Productos Agotados',
                'mensaje': f'{productos_agotados.count()} producto(s) sin stock',
                'productos': productos_agotados
            })
        if productos_criticos.exists():
            alertas.append({
                'tipo': 'warning',
                'titulo': 'Stock Crítico',
                'mensaje': f'{productos_criticos.count()} producto(s) con stock crítico',
                'productos': productos_criticos
            })
        if productos_bajos.exists():
            alertas.append({
                'tipo': 'info',
                'titulo': 'Stock Bajo',
                'mensaje': f'{productos_bajos.count()} producto(s) con stock bajo',
                'productos': productos_bajos
            })
        return alertas
    except Exception as e:
        messages.error(request, f'Error al verificar alertas de stock: {str(e)}')
        return []

@login_required
def detalle_producto(request, pk):
    """Vista de detalle de producto con información completa"""
    producto = get_object_or_404(Producto, pk=pk)

    # Calcular estadísticas básicas (solo ventas activas)
    ventas_mes = VentaDetalle.objects.filter(
        producto=producto,
        venta__fecha__month=timezone.now().month,
        venta__fecha__year=timezone.now().year,
        venta__eliminada=False  # ✅ Excluir ventas eliminadas
    ).aggregate(
        total_vendido=models.Sum('cantidad'),
        total_ventas=models.Count('venta', distinct=True)
    )

    # Obtener últimas ventas del producto (solo activas)
    ultimas_ventas = VentaDetalle.objects.filter(
        producto=producto,
        venta__eliminada=False  # ✅ Excluir ventas eliminadas
    ).select_related('venta').order_by('-venta__fecha')[:5]

    context = {
        'producto': producto,
        'ventas_mes': ventas_mes['total_ventas'] or 0,
        'cantidad_vendida_mes': ventas_mes['total_vendido'] or 0,
        'ultimas_ventas': ultimas_ventas,
    }

    return render(request, 'modules/productos/detalle.html', context)

@login_required
@ratelimit(key='user', rate=getattr(settings, 'RATELIMIT_PRODUCTOS', '50/h'), method='POST', block=True)
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if not request.user.has_perm('gestion.change_producto'):
        messages.error(request, 'No tienes permiso para editar productos.')
        return redirect('gestion:lista_productos')
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Guardar stock anterior para comparar
                    stock_anterior = producto.stock

                    # Guardar el producto
                    producto = form.save(commit=False)

                    # Solo procesar producción si hay cantidad_a_producir explícita
                    # (campo separado del stock, para producir desde materias primas)
                    cantidad_a_producir = form.cleaned_data.get('cantidad_a_producir', 0)
                    if cantidad_a_producir and cantidad_a_producir > 0:
                        # PRODUCCIÓN: Descontar materias primas y aumentar stock
                        ok, faltantes = producto.verificar_stock_materias_primas(cantidad_a_producir)
                        if not ok:
                            faltantes_str = ", ".join([f"{f['materia_prima']} (necesaria: {f['necesaria']} {f['unidad']}, disponible: {f['disponible']})" for f in faltantes])
                            messages.error(request, f"No hay suficiente stock de materias primas para producir {cantidad_a_producir} unidades: {faltantes_str}")
                            raise Exception("Stock de materias primas insuficiente")
                        producto.descontar_materias_primas(cantidad_a_producir, request.user)
                        producto.stock += cantidad_a_producir
                        messages.info(request, f'✅ Producidas {cantidad_a_producir} unidades desde materias primas')

                    # Si solo se cambió el stock manualmente (sin producir)
                    # simplemente guardar el nuevo valor sin descontar nada
                    producto.save()

                    messages.success(request, f'Producto "{producto.nombre}" actualizado exitosamente.')
                return redirect('gestion:lista_productos')
            except Exception as e:
                messages.error(request, f'Error inesperado al actualizar el producto: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
            messages.error(request, 'Error al actualizar el producto. Verifica los datos ingresados.')
    else:
        form = ProductoForm(instance=producto)

    return render(request, 'modules/productos/form.html', {
        'form': form,
        'title': 'Editar Producto',
        'producto': producto
    })

@login_required
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)

    # Verificar permisos
    if not request.user.has_perm('gestion.delete_producto'):
        messages.error(request, 'No tienes permiso para eliminar productos.')
        return redirect('gestion:lista_productos')

    if request.method == 'POST':
        try:
            nombre_producto = producto.nombre
            categoria_producto = producto.get_categoria_display()
            stock_producto = producto.stock

            with transaction.atomic():
                # Verificar si el producto tiene ventas asociadas
                ventas_asociadas = producto.ventadetalle_set.count()
                if ventas_asociadas > 0:
                    messages.warning(
                        request,
                        f'El producto "{nombre_producto}" tiene {ventas_asociadas} venta(s) asociada(s). '
                        'Se eliminará el producto pero se mantendrá el historial de ventas.'
                    )

                # Eliminar el producto
                producto.delete()

                # Mensaje de éxito con información detallada
                messages.success(
                    request,
                    f'Producto "{nombre_producto}" (Categoría: {categoria_producto}, Stock: {stock_producto}) '
                    'eliminado exitosamente.'
                )

            return redirect('gestion:lista_productos')

        except Exception as e:
            messages.error(request, f'Error inesperado al eliminar el producto: {str(e)}')
            return redirect('gestion:lista_productos')

    # Obtener información adicional para mostrar en el template
    ventas_count = producto.ventadetalle_set.count()
    ventas_total = sum(vd.subtotal for vd in producto.ventadetalle_set.all())

    context = {
        'producto': producto,
        'objeto': producto,
        'tipo': 'producto',
        'cancel_url': reverse('gestion:lista_productos'),
        'ventas_asociadas': ventas_count,
        'ventas_total': ventas_total,
        'categoria_display': producto.get_categoria_display(),
        'estado_stock': producto.get_estado_stock(),
    }

    return render(request, 'modules/productos/confirmar_eliminacion_producto.html', context)


@login_required
def eliminar_venta(request, pk):
    """🔒 ELIMINACIÓN SEGURA CON SOFT DELETE - ARQUITECTURA DB PROFESIONAL"""
    venta = get_object_or_404(Venta, pk=pk, eliminada=False)  # Solo ventas activas

    if not request.user.has_perm('gestion.delete_venta'):
        messages.error(request, 'No tienes permiso para eliminar ventas.')
        return redirect('gestion:lista_ventas')

    if request.method == 'POST':
        razon = request.POST.get('razon_eliminacion', '')

        try:
            with transaction.atomic():
                # Guardar información para el mensaje
                monto_venta = venta.total
                productos_restaurados = []

                # Restaurar stock de productos
                detalles = venta.detalles.all()
                for detalle in detalles:
                    producto = detalle.producto
                    producto.stock += detalle.cantidad
                    producto.save()
                    productos_restaurados.append(f"{detalle.producto.nombre} (+{detalle.cantidad})")

                # 🚀 SOFT DELETE - No eliminar, solo marcar
                venta.eliminar_venta(request.user, razon)

                # Mensaje detallado del impacto
                mensaje = 'Venta marcada como eliminada exitosamente. '
                mensaje += 'Se mantiene el historial para auditoría. '
                mensaje += f'Stock restaurado: {", ".join(productos_restaurados)}.'

                messages.success(request, mensaje)

                # Auditoría: registrar acción
                # LogVenta.objects.create(usuario=request.user, accion='eliminar', venta_id=pk, monto=monto_venta)

            return redirect('gestion:lista_ventas')
        except Exception as e:
            messages.error(request, f'Error inesperado al eliminar la venta: {str(e)}')

    # Calcular impacto antes de eliminar (para mostrar en confirmación)
    productos_afectados = []
    for detalle in venta.detalles.all():
        productos_afectados.append({
            'nombre': detalle.producto.nombre,
            'cantidad': detalle.cantidad,
            'stock_actual': detalle.producto.stock,
            'stock_futuro': detalle.producto.stock + detalle.cantidad
        })

    context = {
        'venta': venta,
        'objeto': venta,
        'tipo': 'venta',
        'monto_impacto': venta.total,
        'productos_afectados': productos_afectados,
        'cancel_url': reverse('gestion:lista_ventas')
    }
    return render(request, 'modules/ventas/confirmar_eliminacion_venta.html', context)

@login_required
def detalle_venta(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    return render(request, 'modules/ventas/detalle_venta.html', {'venta': venta})

@login_required
def index(request):
    if request.user.is_authenticated:
        return redirect('gestion:panel_control')
    else:
        return redirect('login')

@login_required
def exportar_productos(request):
    if not request.user.has_perm('gestion.export_producto'):
        messages.error(request, 'No tienes permiso para exportar productos.')
        return redirect('gestion:lista_productos')
    try:
        producto_resource = ProductoResource()
        dataset = producto_resource.export()
        response = HttpResponse(dataset.xlsx, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="productos_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        # Auditoría: registrar acción
        # LogProducto.objects.create(usuario=request.user, accion='exportar', descripcion='Exportación de productos a Excel')
        return response
    except Exception as e:
        messages.error(request, f'Error inesperado al exportar productos: {str(e)}')
        return redirect('gestion:lista_productos')

@login_required
def exportar_ventas(request):
    if not request.user.has_perm('gestion.export_venta'):
        messages.error(request, 'No tienes permiso para exportar ventas.')
        return redirect('gestion:lista_ventas')
    try:
        venta_resource = VentaResource()
        dataset = venta_resource.export()
        response = HttpResponse(dataset.xlsx, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="ventas_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        # Auditoría: registrar acción
        # LogVenta.objects.create(usuario=request.user, accion='exportar', descripcion='Exportación de ventas a Excel')
        return response
    except Exception as e:
        messages.error(request, f'Error inesperado al exportar ventas: {str(e)}')
        return redirect('gestion:lista_ventas')

# ==================== VISTAS MATERIAS PRIMAS ====================

@login_required
def lista_materias_primas(request):
    """Vista para listar materias primas con filtros"""
    materias_primas = MateriaPrima.objects.filter(activo=True)
    query = request.GET.get('q')
    proveedor = request.GET.get('proveedor')
    estado_stock = request.GET.get('estado_stock')

    if query:
        materias_primas = materias_primas.filter(
            Q(nombre__icontains=query) |
            Q(descripcion__icontains=query) |
            Q(proveedor__icontains=query)
        )

    if proveedor:
        materias_primas = materias_primas.filter(proveedor__icontains=proveedor)

    if estado_stock:
        if estado_stock == 'agotado':
            materias_primas = materias_primas.filter(stock_actual=0)
        elif estado_stock == 'bajo':
            materias_primas = materias_primas.filter(stock_actual__gt=0, stock_actual__lte=F('stock_minimo'))
        elif estado_stock == 'normal':
            materias_primas = materias_primas.filter(stock_actual__gt=F('stock_minimo'))

    # Estadísticas para KPIs
    total_materias = MateriaPrima.objects.filter(activo=True).count()
    con_stock = MateriaPrima.objects.filter(activo=True, stock_actual__gt=0).count()
    stock_bajo = MateriaPrima.objects.filter(
        activo=True,
        stock_actual__gt=0,
        stock_actual__lte=F('stock_minimo')
    ).count()
    valor_total = MateriaPrima.objects.filter(activo=True).aggregate(
        total=Sum(F('stock_actual') * F('costo_unitario'))
    )['total'] or 0

    stats = {
        'con_stock': con_stock,
        'stock_bajo': stock_bajo,
        'valor_total': valor_total
    }

    # Obtener proveedores únicos para el filtro
    proveedores = MateriaPrima.objects.filter(activo=True).values_list('proveedor', flat=True).distinct().exclude(proveedor__isnull=True).exclude(proveedor='')

    context = {
        'materias_primas': materias_primas,
        'proveedores': proveedores,
        'stats': stats,
        'query': query or '',
        'proveedor_seleccionado': proveedor or '',
        'estado_stock_seleccionado': estado_stock or '',
    }

    return render(request, 'gestion/materias_primas/lista_simple.html', context)

@login_required
def lista_inventario(request):
    """Vista de inventario optimizada - usa InventarioService para KPIs inteligentes"""
    try:
        from gestion.services.inventario_service import InventarioService

        # Inicializar servicio de inventario
        service = InventarioService()

        # Obtener KPIs inteligentes del servicio
        kpis = service.get_kpis_inventario()

        # Reutilizar lógica existente con paginación
        materias_primas = MateriaPrima.objects.filter(activo=True).order_by('nombre')
        query = request.GET.get('q', '')
        proveedor_seleccionado = request.GET.get('proveedor', '')
        estado_stock = request.GET.get('estado_stock', '')

        # Aplicar filtros (lógica reutilizada)
        if query:
            materias_primas = materias_primas.filter(
                Q(nombre__icontains=query) |
                Q(descripcion__icontains=query) |
                Q(proveedor__icontains=query)
            )

        if proveedor_seleccionado:
            materias_primas = materias_primas.filter(proveedor__icontains=proveedor_seleccionado)

        if estado_stock:
            if estado_stock == 'agotado':
                materias_primas = materias_primas.filter(stock_actual=0)
            elif estado_stock == 'bajo':
                materias_primas = materias_primas.filter(
                    stock_actual__gt=0,
                    stock_actual__lte=F('stock_minimo')
                )
            elif estado_stock == 'normal':
                materias_primas = materias_primas.filter(stock_actual__gt=F('stock_minimo'))

        # Proveedores únicos para filtros
        all_materias = MateriaPrima.objects.filter(activo=True)
        proveedores = (all_materias.values_list('proveedor', flat=True)
                      .distinct()
                      .exclude(proveedor__isnull=True)
                      .exclude(proveedor='')
                      .order_by('proveedor'))

        total_proveedores = len(set(proveedores))

        # Paginación eficiente
        from django.core.paginator import Paginator
        paginator = Paginator(materias_primas, 25)
        page_number = request.GET.get('page', 1)
        materias_paginadas = paginator.get_page(page_number)

        context = {
            'materias_primas': materias_paginadas,
            'proveedores': proveedores,
            'total_proveedores': total_proveedores,
            # KPIs inteligentes del servicio
            'kpis': kpis,
            # KPIs legacy para compatibilidad
            'total_materias': all_materias.count(),
            'con_stock': all_materias.filter(stock_actual__gt=0).count(),
            'stock_bajo': kpis['stock_critico']['cantidad'],
            'stock_critico': kpis['stock_critico']['cantidad'],
            'valor_total': kpis['valor_total']['valor'],
        }

        return render(request, 'modules/inventario/lista_inventario.html', context)

    except Exception as e:
        messages.error(request, f'Error al cargar inventario: {str(e)}')
        return redirect('gestion:panel_control')

@login_required
def crear_materia_prima(request):
    """Vista para crear nueva materia prima"""
    if not request.user.has_perm('gestion.add_materiaprima'):
        messages.error(request, 'No tienes permiso para crear materias primas.')
        return redirect('gestion:lista_inventario')
    if request.method == 'POST':
        form = MateriaPrimaForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    materia_prima = form.save()
                    # Registrar movimiento inicial si hay stock
                    if materia_prima.stock_actual > 0:
                        MovimientoMateriaPrima.objects.create(
                            materia_prima=materia_prima,
                            tipo_movimiento='entrada',
                            cantidad=materia_prima.stock_actual,
                            cantidad_anterior=0,
                            cantidad_nueva=materia_prima.stock_actual,
                            motivo='Stock inicial',
                            usuario=request.user
                        )
                    # Auditoría: registrar acción
                    # LogMateriaPrima.objects.create(usuario=request.user, accion='crear', materia_prima=materia_prima)
                messages.success(request, f'Materia prima "{materia_prima.nombre}" creada exitosamente.')
                return redirect('gestion:lista_inventario')
            except Exception as e:
                messages.error(request, f'Error inesperado al crear la materia prima: {str(e)}')
        else:
            messages.error(request, 'Error al crear la materia prima. Verifica los datos.')
    else:
        form = MateriaPrimaForm()
    return render(request, 'modules/materias_primas/materias_primas/crear.html', {
        'form': form,
        'titulo': 'Crear Materia Prima'
    })

@login_required
def editar_materia_prima(request, pk):
    """Vista para editar materia prima"""
    materia_prima = get_object_or_404(MateriaPrima, pk=pk)
    if not request.user.has_perm('gestion.change_materiaprima'):
        messages.error(request, 'No tienes permiso para editar materias primas.')
        return redirect('gestion:lista_inventario')
    stock_anterior = materia_prima.stock_actual
    if request.method == 'POST':
        form = MateriaPrimaForm(request.POST, instance=materia_prima)
        if form.is_valid():
            try:
                with transaction.atomic():
                    materia_prima = form.save()
                    # Registrar movimiento si cambió el stock
                    if stock_anterior != materia_prima.stock_actual:
                        diferencia = materia_prima.stock_actual - stock_anterior
                        tipo_mov = 'entrada' if diferencia > 0 else 'salida'
                        MovimientoMateriaPrima.objects.create(
                            materia_prima=materia_prima,
                            tipo_movimiento='ajuste',
                            cantidad=abs(diferencia),
                            cantidad_anterior=stock_anterior,
                            cantidad_nueva=materia_prima.stock_actual,
                            motivo=f'Ajuste manual - {tipo_mov}',
                            usuario=request.user
                        )
                    # Auditoría: registrar acción
                    # LogMateriaPrima.objects.create(usuario=request.user, accion='editar', materia_prima=materia_prima)
                messages.success(request, f'Materia prima "{materia_prima.nombre}" actualizada exitosamente.')
                return redirect('gestion:lista_inventario')
            except Exception as e:
                messages.error(request, f'Error inesperado al actualizar la materia prima: {str(e)}')
        else:
            messages.error(request, 'Error al actualizar la materia prima. Verifica los datos.')
    else:
        form = MateriaPrimaForm(instance=materia_prima)
    return render(request, 'modules/materias_primas/materias_primas/form.html', {
        'form': form,
        'titulo': 'Editar Materia Prima',
        'materia_prima': materia_prima
    })

@login_required
def detalle_materia_prima(request, pk):
    """Vista detalle de materia prima con movimientos, lotes FIFO, permisos y errores"""
    if not request.user.has_perm('gestion.view_materiaprima'):
        messages.error(request, 'No tienes permiso para ver detalles de materias primas.')
        return redirect('gestion:lista_inventario')
    try:
        materia_prima = get_object_or_404(MateriaPrima, pk=pk)
        movimientos = materia_prima.movimientos.all()[:20]  # Últimos 20 movimientos
        productos_relacionados = ProductoMateriaPrima.objects.filter(materia_prima=materia_prima)
        # Importar el modelo de lotes y obtener los lotes FIFO de esta materia prima
        from .models import LoteMateriaPrima
        lotes = LoteMateriaPrima.objects.filter(materia_prima=materia_prima).order_by('fecha_entrada', 'id')
        context = {
            'materia_prima': materia_prima,
            'movimientos': movimientos,
            'productos_relacionados': productos_relacionados,
            'lotes': lotes,
        }
        # Auditoría: registrar acción
        # LogMateriaPrima.objects.create(usuario=request.user, accion='ver', materia_prima=materia_prima, descripcion='Detalle de materia prima')
        return render(request, 'modules/materias_primas/materias_primas/detalle.html', context)
    except Exception as e:
        messages.error(request, f'Error inesperado al mostrar detalle: {str(e)}')
        return redirect('gestion:lista_inventario')

@login_required
def movimiento_materia_prima(request, pk):
    """Vista para registrar movimiento de materia prima con permisos, errores y auditoría"""
    if not request.user.has_perm('gestion.change_materiaprima'):
        messages.error(request, 'No tienes permiso para registrar movimientos de materias primas.')
        return redirect('gestion:lista_inventario')
    try:
        materia_prima = get_object_or_404(MateriaPrima, pk=pk)
        if request.method == 'POST':
            form = MovimientoMateriaPrimaForm(request.POST)
            if form.is_valid():
                movimiento = form.save(commit=False)
                movimiento.usuario = request.user
                movimiento.cantidad_anterior = materia_prima.stock_actual
                # Actualizar stock según tipo de movimiento
                if movimiento.tipo_movimiento in ['entrada', 'devolucion']:
                    materia_prima.stock_actual += movimiento.cantidad
                elif movimiento.tipo_movimiento in ['salida', 'produccion']:
                    if materia_prima.stock_actual >= movimiento.cantidad:
                        # FIFO: descontar de lotes
                        from .models import LoteMateriaPrima
                        cantidad_restante = float(movimiento.cantidad)
                        lotes = LoteMateriaPrima.objects.filter(materia_prima=materia_prima, cantidad_disponible__gt=0).order_by('fecha_entrada', 'id')
                        for lote in lotes:
                            if cantidad_restante <= 0:
                                break
                            disponible = float(lote.cantidad_disponible)
                            a_consumir = min(disponible, cantidad_restante)
                            lote.cantidad_disponible = disponible - a_consumir
                            if lote.cantidad_disponible == 0:
                                lote.fecha_consumo = timezone.now().date()
                            lote.save()
                            cantidad_restante -= a_consumir
                        materia_prima.stock_actual -= movimiento.cantidad
                    else:
                        messages.error(request, 'No hay suficiente stock disponible.')
                        return render(request, 'modules/materias_primas/materias_primas/movimiento.html', {
                            'form': form,
                            'materia_prima': materia_prima
                        })
                elif movimiento.tipo_movimiento == 'ajuste':
                    materia_prima.stock_actual = movimiento.cantidad
                movimiento.cantidad_nueva = materia_prima.stock_actual
                # Guardar ambos objetos
                with transaction.atomic():
                    materia_prima.save()
                    movimiento.save()
                # Auditoría: registrar acción
                # LogMateriaPrima.objects.create(usuario=request.user, accion='movimiento', materia_prima=materia_prima, descripcion=f'Movimiento: {movimiento.tipo_movimiento}')
                messages.success(request, f'Movimiento registrado exitosamente. Nuevo stock: {materia_prima.stock_actual}')
                return redirect('gestion:detalle_materia_prima', pk=pk)
            else:
                messages.error(request, 'Error al registrar el movimiento. Verifica los datos.')
        else:
            form = MovimientoMateriaPrimaForm(initial={'materia_prima': materia_prima})
        return render(request, 'modules/materias_primas/materias_primas/movimiento.html', {
            'form': form,
            'materia_prima': materia_prima
        })
    except Exception as e:
        messages.error(request, f'Error inesperado al registrar movimiento: {str(e)}')
        return redirect('gestion:detalle_materia_prima', pk=pk)


# ====== VISTAS DE REPORTES AVANZADOS ======

@login_required
def reporte_stock_materias_primas(request):
    """Reporte detallado de stock de materias primas. Incluye clasificación por estado y valores totales."""
    if not request.user.has_perm('gestion.view_reporte'):
        messages.error(request, 'No tienes permiso para ver reportes de stock de materias primas.')
        return redirect('gestion:panel_control')
    try:
        materias_primas = MateriaPrima.objects.filter(activo=True)
        # Clasificación de materias primas según estado de stock
        stock_critico = materias_primas.filter(stock_actual__lte=models.F('stock_minimo'))
        stock_bajo = materias_primas.filter(
            stock_actual__gt=models.F('stock_minimo'),
            stock_actual__lte=models.F('stock_minimo') * 2
        )
        stock_normal = materias_primas.exclude(
            id__in=list(stock_critico.values_list('id', flat=True)) +
                   list(stock_bajo.values_list('id', flat=True))
        )
        # Cálculo de valores totales y críticos
        valor_total = sum([mp.valor_total_stock for mp in materias_primas])
        valor_critico = sum([mp.valor_total_stock for mp in stock_critico])
        context = {
            'materias_primas': materias_primas,
            'stock_critico': stock_critico,
            'stock_bajo': stock_bajo,
            'stock_normal': stock_normal,
            'valor_total': valor_total,
            'valor_critico': valor_critico,
            'total_materias': materias_primas.count(),
        }
        # Auditoría: acción de consulta de reporte de stock (descomentar si se implementa logging)
        # LogReporte.objects.create(usuario=request.user, accion='ver', descripcion='Reporte de stock de materias primas')
        return render(request, 'gestion/reportes/stock_materias_primas.html', context)
    except Exception as e:
        messages.error(request, f'Error inesperado al generar el reporte de stock: {str(e)}')
        return redirect('gestion:panel_control')

@login_required
def reporte_costos_produccion(request):
    """Reporte de costos de producción por producto. Analiza margen y porcentaje de ganancia."""
    if not request.user.has_perm('gestion.view_reporte'):
        messages.error(request, 'No tienes permiso para ver reportes de costos de producción.')
        return redirect('gestion:panel_control')
    try:
        # Analiza productos con receta y calcula margen de ganancia
        productos_con_receta = Producto.objects.filter(recetas__isnull=False).distinct()
        productos_analisis = []
        for producto in productos_con_receta:
            costo_materias = producto.costo_materias_primas
            margen_ganancia = producto.precio - costo_materias
            porcentaje_margen = (margen_ganancia / producto.precio * 100) if producto.precio > 0 else 0
            productos_analisis.append({
                'producto': producto,
                'costo_materias': costo_materias,
                'precio_venta': producto.precio,
                'margen_ganancia': margen_ganancia,
                'porcentaje_margen': porcentaje_margen,
            })
        productos_analisis.sort(key=lambda x: x['porcentaje_margen'], reverse=True)
        context = {
            'productos_analisis': productos_analisis,
            'total_productos': len(productos_analisis),
        }
        return render(request, 'gestion/reportes/costos_produccion.html', context)
    except Exception as e:
        messages.error(request, f'Error inesperado al generar el reporte de costos: {str(e)}')
        return redirect('gestion:panel_control')

# ====== EXPORTACIÓN MEJORADA ======

@login_required
def exportar_materias_primas_excel(request):
    if not request.user.has_perm('gestion.export_materiaprima'):
        messages.error(request, 'No tienes permiso para exportar materias primas.')
        return redirect('gestion:lista_inventario')
    try:
        # Exporta materias primas a Excel con formato y estilos
        from datetime import datetime

        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Materias Primas"
        headers = ['Nombre', 'Unidad', 'Stock Actual', 'Stock Mínimo', 'Costo Unitario',
                   'Valor Total', 'Proveedor', 'Estado Stock']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        materias_primas = MateriaPrima.objects.filter(activo=True)
        for row, mp in enumerate(materias_primas, 2):
            ws.cell(row=row, column=1, value=mp.nombre)
            ws.cell(row=row, column=2, value=mp.get_unidad_medida_display())
            ws.cell(row=row, column=3, value=float(mp.stock_actual))
            ws.cell(row=row, column=4, value=float(mp.stock_minimo))
            ws.cell(row=row, column=5, value=float(mp.costo_unitario))
            ws.cell(row=row, column=6, value=float(mp.valor_total_stock))
            ws.cell(row=row, column=7, value=mp.proveedor or '')
            ws.cell(row=row, column=8, value='Crítico' if mp.necesita_restock else 'Normal')
        # Ajusta ancho de columnas automáticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="materias_primas_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        # Auditoría: acción de exportación de materias primas (descomentar si se implementa logging)
        # LogMateriaPrima.objects.create(usuario=request.user, accion='exportar', descripcion='Exportación de materias primas a Excel')
        return response
    except Exception as e:
        messages.error(request, f'Error inesperado al exportar materias primas: {str(e)}')
        return redirect('gestion:lista_inventario')

@login_required
def exportar_reporte_pdf(request, tipo_reporte):
    if not request.user.has_perm('gestion.export_reporte'):
        messages.error(request, 'No tienes permiso para exportar reportes.')
        return redirect('gestion:reportes')
    try:
        from io import BytesIO

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        if tipo_reporte == 'stock_materias':
            title = Paragraph("Reporte de Stock - Materias Primas", title_style)
            elements.append(title)
            elements.append(Spacer(1, 12))
            # Datos
            materias_primas = MateriaPrima.objects.filter(activo=True)
            data = [['Nombre', 'Stock Actual', 'Stock Mínimo', 'Estado']]
            for mp in materias_primas:
                estado = 'CRÍTICO' if mp.necesita_restock else 'Normal'
                data.append([
                    mp.nombre,
                    f"{mp.stock_actual} {mp.get_unidad_medida_display()}",
                    f"{mp.stock_minimo} {mp.get_unidad_medida_display()}",
                    estado
                ])
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
        elif tipo_reporte == 'costos_produccion':
            title = Paragraph("Reporte de Costos de Producción", title_style)
            elements.append(title)
            elements.append(Spacer(1, 12))
            productos_con_receta = Producto.objects.filter(recetas__isnull=False).distinct()
            data = [['Producto', 'Costo Materias', 'Precio Venta', 'Margen', '% Margen']]
            for producto in productos_con_receta:
                costo_materias = producto.costo_materias_primas
                margen = producto.precio - costo_materias
                porcentaje = (margen / producto.precio * 100) if producto.precio > 0 else 0
                data.append([
                    producto.nombre,
                    f"${costo_materias:.2f}",
                    f"${producto.precio:.2f}",
                    f"${margen:.2f}",
                    f"{porcentaje:.1f}%"
                ])
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_{tipo_reporte}_{datetime.now().strftime("%Y%m%d")}.pdf"'
        # Auditoría: acción de exportación de reporte PDF (descomentar si se implementa logging)
        # LogReporte.objects.create(usuario=request.user, accion='exportar', descripcion=f'Exportación de reporte PDF: {tipo_reporte}')
        return response
    except Exception as e:
        messages.error(request, f'Error inesperado al exportar el reporte PDF: {str(e)}')
        return redirect('gestion:reportes')

# ====== API ENDPOINTS ======

@login_required
def api_verificar_stock_producto(request, producto_id):
    """API para verificar stock de producto y materias primas. Devuelve estado y faltantes."""
    if not request.user.has_perm('gestion.view_producto'):
        return JsonResponse({'success': False, 'error': 'No tienes permiso para consultar stock.'}, status=403)
    try:
        producto = Producto.objects.get(id=producto_id)
        cantidad = int(request.GET.get('cantidad', 1))
        # Verifica stock del producto
        stock_producto_ok = producto.stock >= cantidad
        # Verifica stock de materias primas si el producto tiene receta
        stock_materias_ok = True
        faltantes = []
        if producto.tiene_receta:
            stock_materias_ok, faltantes = producto.verificar_stock_materias_primas(cantidad)
        # Auditoría: consulta API de stock (descomentar si se implementa logging)
        # LogProducto.objects.create(usuario=request.user, accion='api_verificar_stock', producto=producto, descripcion=f'Consulta de stock para {cantidad}')
        return JsonResponse({
            'success': True,
            'stock_producto_ok': stock_producto_ok,
            'stock_disponible': producto.stock,
            'stock_materias_ok': stock_materias_ok,
            'faltantes': faltantes,
            'tiene_receta': producto.tiene_receta
        })
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Producto no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_costo_receta(request, pk):
    """API endpoint para obtener el costo total de una receta."""
    try:
        receta = get_object_or_404(Receta, pk=pk)
        costo_total = receta.costo_total()

        # También devolver información detallada de ingredientes para debug
        ingredientes = []
        for ingrediente in receta.recetamateriaprima_set.all():
            costo_ingrediente = ingrediente.cantidad * ingrediente.materia_prima.costo_unitario
            ingredientes.append({
                'nombre': ingrediente.materia_prima.nombre,
                'cantidad': float(ingrediente.cantidad),
                'costo_unitario': float(ingrediente.materia_prima.costo_unitario),
                'costo_total': float(costo_ingrediente)
            })

        return JsonResponse({
            'success': True,
            'costo_total': float(costo_total),
            'ingredientes': ingredientes,
            'nombre_receta': receta.nombre
        })
    except Receta.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Receta no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# ==================== VISTA DEMO COMPONENTES ====================
@login_required
def demo_componentes(request):
    """Vista demo para mostrar los nuevos componentes Lino."""
    return render(request, 'gestion/demo_componentes.html', {
        'titulo': 'Demo - Componentes Lino'
    })

# ===== VISTAS LINO V3 =====

@login_required
def reportes_lino(request):
    """Vista migrada de reportes usando el sistema de diseño Lino"""
    try:
        from datetime import datetime, timedelta
        from decimal import Decimal


        hoy = datetime.now().date()

        # Obtener rango de fechas desde request o usar mes actual por defecto
        fecha_desde_str = request.GET.get('fecha_desde', '')
        fecha_hasta_str = request.GET.get('fecha_hasta', '')

        if fecha_desde_str and fecha_hasta_str:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d').date()
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d').date()
        else:
            # Por defecto: mes actual
            fecha_desde = hoy.replace(day=1)
            fecha_hasta = hoy

        # Calcular mes anterior para comparación
        dias_periodo = (fecha_hasta - fecha_desde).days
        fecha_desde_anterior = fecha_desde - timedelta(days=dias_periodo + 1)
        fecha_hasta_anterior = fecha_desde - timedelta(days=1)

        # Cálculos del período actual
        ventas = Venta.objects.filter(fecha__range=[fecha_desde, fecha_hasta])
        compras = Compra.objects.filter(fecha_compra__range=[fecha_desde, fecha_hasta])

        # Cálculos del período anterior
        ventas_anterior = Venta.objects.filter(fecha__range=[fecha_desde_anterior, fecha_hasta_anterior])
        compras_anterior = Compra.objects.filter(fecha_compra__range=[fecha_desde_anterior, fecha_hasta_anterior])

        # Productos y materias primas (no filtrados por fecha)
        productos = Producto.objects.all()
        materias_primas = MateriaPrima.objects.all()

        # === PERÍODO ACTUAL ===
        ingresos_totales = sum(Decimal(str(venta.total)) for venta in ventas)
        # Calcular gastos desde detalles o precio_mayoreo legacy
        gastos_totales = Decimal('0')
        for compra in compras:
            if compra.precio_mayoreo:
                gastos_totales += Decimal(str(compra.precio_mayoreo))
            else:
                # Calcular desde detalles
                for detalle in compra.detalles.all():
                    gastos_totales += Decimal(str(detalle.precio_unitario)) * Decimal(str(detalle.cantidad))

        ganancia_neta = ingresos_totales - gastos_totales
        total_ventas = ventas.count()

        # === PERÍODO ANTERIOR ===
        ingresos_anterior = sum(Decimal(str(venta.total)) for venta in ventas_anterior)
        # Calcular gastos desde detalles o precio_mayoreo legacy
        gastos_anterior = Decimal('0')
        for compra in compras_anterior:
            if compra.precio_mayoreo:
                gastos_anterior += Decimal(str(compra.precio_mayoreo))
            else:
                # Calcular desde detalles
                for detalle in compra.detalles.all():
                    gastos_anterior += Decimal(str(detalle.precio_unitario)) * Decimal(str(detalle.cantidad))

        ganancia_anterior = ingresos_anterior - gastos_anterior
        total_ventas_anterior = ventas_anterior.count()

        # === CALCULAR VARIACIONES ===
        def calcular_variacion(actual, anterior):
            if anterior > 0:
                return float(((actual - anterior) / anterior) * 100)
            elif actual > 0:
                return 100.0  # Crecimiento del 100% si no había datos anteriores
            else:
                return 0.0

        variacion_ingresos = calcular_variacion(ingresos_totales, ingresos_anterior)
        variacion_gastos = calcular_variacion(gastos_totales, gastos_anterior)
        variacion_ganancia = calcular_variacion(ganancia_neta, ganancia_anterior)
        variacion_ventas = calcular_variacion(total_ventas, total_ventas_anterior)

        # Calcular margen y ROI
        margen_porcentaje = float((ganancia_neta / ingresos_totales) * 100) if ingresos_totales > 0 else 0
        roi = float((ganancia_neta / gastos_totales) * 100) if gastos_totales > 0 else 0

        # Métricas adicionales
        ticket_promedio = float(ingresos_totales / total_ventas) if total_ventas > 0 else 0
        ticket_promedio_anterior = float(ingresos_anterior / total_ventas_anterior) if total_ventas_anterior > 0 else 0
        variacion_ticket = calcular_variacion(ticket_promedio, ticket_promedio_anterior)

        # Productos y stock
        productos_criticos = productos.filter(stock__lte=F('stock_minimo')).count()
        valor_inventario = sum((Decimal(str(p.precio or 0)) * Decimal(str(p.stock or 0))) for p in productos)

        # Proveedores activos
        proveedores_activos = len(set(mp.proveedor for mp in materias_primas if mp.proveedor))

        # Alertas
        alertas = []
        if productos_criticos > 0:
            alertas.append({
                'tipo': 'warning',
                'titulo': 'Stock Crítico',
                'descripcion': f'{productos_criticos} productos requieren reposición'
            })

        if ganancia_neta < 0:
            alertas.append({
                'tipo': 'danger',
                'titulo': 'Pérdidas Detectadas',
                'descripcion': 'La ganancia neta es negativa, revisa los costos'
            })

        if variacion_ingresos < -10:
            alertas.append({
                'tipo': 'warning',
                'titulo': 'Caída en Ventas',
                'descripcion': f'Los ingresos bajaron {abs(variacion_ingresos):.1f}% vs período anterior'
            })

        context = {
            # Filtros
            'fecha_desde': fecha_desde.strftime('%Y-%m-%d'),
            'fecha_hasta': fecha_hasta.strftime('%Y-%m-%d'),

            # Período actual
            'ingresos_totales': float(ingresos_totales),
            'gastos_totales': float(gastos_totales),
            'ganancia_neta': float(ganancia_neta),
            'margen_porcentaje': margen_porcentaje,
            'roi': roi,
            'total_ventas': total_ventas,
            'total_compras': compras.count(),
            'ticket_promedio': ticket_promedio,

            # Variaciones vs período anterior
            'variacion_ingresos': variacion_ingresos,
            'variacion_gastos': variacion_gastos,
            'variacion_ganancia': variacion_ganancia,
            'variacion_ventas': variacion_ventas,
            'variacion_ticket': variacion_ticket,

            # Inventario y productos
            'total_productos': productos.count(),
            'productos_criticos': productos_criticos,
            'valor_inventario': float(valor_inventario),
            'proveedores_activos': proveedores_activos,

            # Campos legacy (mantener compatibilidad)
            'margen_bruto': margen_porcentaje,
            'inversion_total': float(gastos_totales),
            'crecimiento_ventas': variacion_ingresos,
            'rotacion_inventario': 'N/A',
            'clientes_recurrentes': 0,
            'productos_top_count': 5,
            'alertas': alertas,
        }

        return render(request, 'modules/reportes/dashboard_enterprise.html', context)

    except Exception as e:
        messages.error(request, f'Error al generar reportes: {str(e)}')
        return redirect('gestion:panel_control')


# ==================== VISTAS DE ANALYTICS Y CONTROL DE RENTABILIDAD ====================

@login_required
def dashboard_rentabilidad(request):
    """
    Dashboard principal de control de rentabilidad con objetivos de negocio.
    Muestra: objetivo de margen, productos críticos, recomendaciones.
    """
    try:
        from django.core.paginator import Paginator

        from gestion.services.rentabilidad_service import RentabilidadService

        # Usar el nuevo servicio de rentabilidad
        service = RentabilidadService()

        # Obtener KPIs principales
        kpis = service.get_kpis_rentabilidad()

        # Obtener análisis detallado de objetivo de margen
        analisis_objetivo = service.get_objetivo_margen_analisis()

        # Obtener todos los productos con su rentabilidad para la tabla
        productos_rentabilidad = service.get_productos_rentabilidad()

        # 📄 PAGINACIÓN DE LA TABLA (15 productos por página)
        paginator = Paginator(productos_rentabilidad, 15)
        page_number = request.GET.get('page', 1)
        productos_paginados = paginator.get_page(page_number)

        # Datos para gráfico de distribución de márgenes
        margenes_labels = [
            'En Pérdida',
            'Crítico (<10%)',
            'Bajo (10-20%)',
            'Aceptable (20-30%)',
            'Bueno (30-40%)',
            'Muy Bueno (40-60%)',
            'Excelente (60-80%)',
            'Premium (80-100%)',
            'Elite (>100%)'
        ]
        margenes_data = [
            len([p for p in productos_rentabilidad if p['en_perdida']]),
            len([p for p in productos_rentabilidad if p['margen'] < 10 and not p['en_perdida']]),
            len([p for p in productos_rentabilidad if 10 <= p['margen'] < 20]),
            len([p for p in productos_rentabilidad if 20 <= p['margen'] < 30]),
            len([p for p in productos_rentabilidad if 30 <= p['margen'] < 40]),
            len([p for p in productos_rentabilidad if 40 <= p['margen'] < 60]),
            len([p for p in productos_rentabilidad if 60 <= p['margen'] < 80]),
            len([p for p in productos_rentabilidad if 80 <= p['margen'] < 100]),
            len([p for p in productos_rentabilidad if p['margen'] >= 100])
        ]

        # Top 10 productos por margen
        top_margenes = sorted(
            [p for p in productos_rentabilidad if not p['en_perdida']],
            key=lambda x: x['margen'],
            reverse=True
        )[:10]
        top_margenes_labels = [p['nombre'][:20] for p in top_margenes]
        top_margenes_data = [p['margen'] for p in top_margenes]

        # ADAPTAR ESTRUCTURA PARA EL TEMPLATE
        # El template espera kpis.objetivo_margen, kpis.rentables.total, etc.
        total_productos = analisis_objetivo['total_productos']

        kpis_adaptados = {
            'objetivo_margen': {
                'meta': analisis_objetivo['meta'],
                'actual': analisis_objetivo['actual'],
                'gap': analisis_objetivo['gap'],
                'progreso': analisis_objetivo['progreso'],
                'alcanzado': analisis_objetivo['alcanzado'],
            },
            'rentables': {
                'porcentaje': kpis['rentables']['porcentaje'],
                'cantidad': kpis['rentables']['cantidad'],
                'total': total_productos,
            },
            'en_perdida': {
                'porcentaje': kpis['en_perdida']['porcentaje'],
                'cantidad': kpis['en_perdida']['cantidad'],
                'total': total_productos,
            },
            'margen_promedio': {
                'valor': kpis['margen_promedio'],
                'ponderado': True,
            },
        }

        # Productos críticos (los que no cumplen objetivo)
        productos_criticos = [p for p in productos_rentabilidad if not p['cumple_objetivo'] and not p['en_perdida']]

        analisis_adaptado = {
            **analisis_objetivo,
            'productos_criticos': productos_criticos[:10],
        }

        context = {
            'kpis': kpis_adaptados,
            'analisis_objetivo': analisis_adaptado,
            'productos_paginados': productos_paginados,
            'margenes_labels': json.dumps(margenes_labels),
            'margenes_data': json.dumps(margenes_data),
            'top_margenes_labels': json.dumps(top_margenes_labels),
            'top_margenes_data': json.dumps(top_margenes_data),
        }

        return render(request, 'gestion/dashboard_rentabilidad_v3.html', context)

    except Exception as e:
        messages.error(request, f'Error al cargar dashboard de rentabilidad: {str(e)}')
        logger.error("Error en dashboard_rentabilidad: %s", e, exc_info=True)
        return redirect('gestion:panel_control')


@login_required
def detalle_rentabilidad_producto(request, producto_id):
    """
    Vista detallada de rentabilidad de un producto específico
    """
    try:
        producto = get_object_or_404(Producto, id=producto_id)
        analytics = AnalyticsRentabilidad()

        # Datos específicos del producto
        evolucion = analytics.get_evolucion_costos(producto_id=producto_id)
        productos_rentabilidad = analytics.get_productos_rentabilidad()

        # Encontrar los datos del producto específico
        producto_data = next((p for p in productos_rentabilidad if p['producto'].id == producto_id), None)

        # Histórico de ventas (últimos 30 días)
        desde = timezone.now().date() - timedelta(days=30)
        ventas_historico = VentaDetalle.objects.filter(
            producto=producto,
            venta__fecha__date__gte=desde
        ).order_by('venta__fecha')

        # Preparar datos para gráfico de evolución
        fechas = []
        precios = []
        cantidades = []

        for venta in ventas_historico:
            fechas.append(venta.venta.fecha.strftime('%Y-%m-%d'))
            precios.append(float(venta.precio_unitario))
            cantidades.append(venta.cantidad)

        context = {
            'producto': producto,
            'producto_data': producto_data,
            'evolucion': evolucion[0] if evolucion else None,
            'ventas_historico': ventas_historico,
            'fechas_json': json.dumps(fechas),
            'precios_json': json.dumps(precios),
            'cantidades_json': json.dumps(cantidades)
        }

        return render(request, 'gestion/detalle_rentabilidad_producto.html', context)

    except Exception as e:
        messages.error(request, f'Error al cargar detalle de rentabilidad: {str(e)}')
        return redirect('gestion:dashboard_rentabilidad')


@login_required
def alertas_rentabilidad_ajax(request):
    """
    Vista AJAX para obtener alertas de rentabilidad en tiempo real
    """
    try:
        analytics = AnalyticsRentabilidad()
        alertas = analytics.get_alertas_rentabilidad()

        return JsonResponse({
            'success': True,
            'alertas': alertas,
            'total': len(alertas)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def recomendaciones_precios_ajax(request):
    """
    Vista AJAX para obtener recomendaciones de ajuste de precios
    """
    try:
        analytics = AnalyticsRentabilidad()
        recomendaciones = analytics.get_recomendaciones_precios()

        # Serializar datos para JSON
        recomendaciones_json = []
        for rec in recomendaciones:
            recomendaciones_json.append({
                'producto_id': rec['producto'].id,
                'producto_nombre': rec['producto'].nombre,
                'tipo': rec['tipo'],
                'problema': rec['problema'],
                'precio_actual': rec['precio_actual'],
                'precio_sugerido': rec['precio_sugerido'],
                'incremento_porcentaje': rec['incremento_porcentaje'],
                'justificacion': rec['justificacion']
            })

        return JsonResponse({
            'success': True,
            'recomendaciones': recomendaciones_json,
            'total': len(recomendaciones_json)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def aplicar_precio_sugerido(request, producto_id):
    """
    Vista para aplicar un precio sugerido a un producto
    """
    if request.method == 'POST':
        try:
            with transaction.atomic():
                producto = get_object_or_404(Producto, id=producto_id)
                nuevo_precio = float(request.POST.get('nuevo_precio'))
                precio_anterior = producto.precio

                # Validar precio
                if nuevo_precio <= 0:
                    return JsonResponse({
                        'success': False,
                        'error': 'El precio debe ser mayor a 0'
                    })

                # Actualizar precio
                producto.precio = nuevo_precio
                producto.save()

                # Log de la operación
                log_business_operation(
                    user=request.user,
                    operation='ajuste_precio',
                    details=f'Producto: {producto.nombre} - Precio anterior: ${precio_anterior} - Precio nuevo: ${nuevo_precio}',
                    related_objects={'producto': producto}
                )

                return JsonResponse({
                    'success': True,
                    'message': f'Precio actualizado exitosamente. ${precio_anterior} → ${nuevo_precio}'
                })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al actualizar precio: {str(e)}'
            })

    return JsonResponse({'success': False, 'error': 'Método no permitido'})


# ==================== VISTAS V3 - FORMULARIOS ====================
# crear_venta_v3, crear_compra_v3, crear_receta_v3 eliminadas — sin URL, sin referencias.
# Canónicas: views_ventas.crear_venta, views_compras.crear_compra, views_recetas.crear_receta


# ==================== IMPORTAR FUNCIONES API ====================


# ============================================
# FASE 3: SISTEMA DE ALERTAS UI
# ============================================

@login_required
def alertas_count_api(request):
    """
    API: Retorna count de alertas no leídas para el badge del navbar
    
    Response:
        {
            "count": 5
        }
    """
    from .services.alertas_service import AlertasService

    service = AlertasService()
    count = service.get_alertas_count(usuario=request.user, solo_no_leidas=True)

    return JsonResponse({'count': count})


@login_required
def alertas_no_leidas_api(request):
    """
    API: Retorna últimas 5 alertas no leídas para el slide-in panel
    
    Response:
        {
            "alertas": [
                {
                    "id": 1,
                    "tipo": "stock_bajo",
                    "nivel": "danger",
                    "titulo": "Stock Crítico",
                    "mensaje": "Almendras: solo quedan 0.5kg",
                    "fecha": "04/11/2025 19:00",
                    "icono": "bi-box-seam"
                },
                ...
            ]
        }
    """
    from .services.alertas_service import AlertasService

    service = AlertasService()
    alertas = service.get_alertas_usuario(
        usuario=request.user,
        solo_no_leidas=True,
        limit=5
    )

    # Serializar alertas a JSON
    data = [{
        'id': a.id,
        'tipo': a.tipo,
        'nivel': a.nivel,
        'titulo': a.titulo,
        'mensaje': a.mensaje,
        'fecha': a.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
        'icono': a.get_icono() if hasattr(a, 'get_icono') else 'bi-info-circle',
    } for a in alertas]

    return JsonResponse({'alertas': data})


@require_POST
@login_required
def marcar_alerta_leida(request, alerta_id):
    """
    API: Marca una alerta como leída vía AJAX
    
    Method: POST
    
    Response:
        {
            "success": true
        }
        
        o
        
        {
            "success": false,
            "error": "Alerta no encontrada"
        }
    """
    from django.utils import timezone

    from .models import Alerta

    try:
        alerta = Alerta.objects.get(id=alerta_id, usuario=request.user)
        alerta.leida = True
        alerta.fecha_lectura = timezone.now()
        alerta.save()

        return JsonResponse({'success': True})

    except Alerta.DoesNotExist:
        return JsonResponse(
            {'success': False, 'error': 'Alerta no encontrada'},
            status=404
        )


@login_required
def alertas_lista(request):
    """
    Vista: Página completa de lista de alertas con filtros y paginación
    
    GET params:
        - tipo: filtrar por tipo de alerta
        - nivel: filtrar por nivel (danger, warning, info)
        - no_leidas: boolean, mostrar solo no leídas
        - page: número de página
    """
    from django.core.paginator import Paginator

    from .services.alertas_service import AlertasService

    service = AlertasService()

    # Obtener filtros de GET
    tipo = request.GET.get('tipo', None)
    nivel = request.GET.get('nivel', None)
    solo_no_leidas = request.GET.get('no_leidas', 'false').lower() == 'true'

    # Obtener alertas filtradas
    alertas = service.get_alertas_usuario(
        usuario=request.user,
        tipo=tipo if tipo else None,
        nivel=nivel if nivel else None,
        solo_no_leidas=solo_no_leidas
    )

    # Paginación (20 por página)
    paginator = Paginator(alertas, 20)
    page_number = request.GET.get('page', 1)
    alertas_page = paginator.get_page(page_number)

    # Opciones para filtros
    tipos_disponibles = [
        ('stock_bajo', 'Stock Bajo'),
        ('vencimiento', 'Próximo a Vencer'),
        ('precio_cambio', 'Cambio de Precio'),
        ('stock_critico', 'Stock Crítico'),
    ]

    niveles_disponibles = [
        ('danger', 'Crítico'),
        ('warning', 'Advertencia'),
        ('info', 'Información'),
        ('success', 'Éxito'),
    ]

    context = {
        'title': 'Gestión de Alertas',
        'alertas': alertas_page,
        'tipo_actual': tipo,
        'nivel_actual': nivel,
        'solo_no_leidas': solo_no_leidas,
        'tipos_disponibles': tipos_disponibles,
        'niveles_disponibles': niveles_disponibles,
    }

    return render(request, 'gestion/alertas_lista.html', context)


@login_required
def configuracion_negocio(request):
    """
    Vista para configurar objetivos y parámetros del negocio.
    Permite al dueño establecer: margen objetivo, rotación, cobertura.
    """
    from decimal import Decimal

    from gestion.models import ConfiguracionCostos

    # Obtener o crear configuración
    config = ConfiguracionCostos.objects.first()
    if not config:
        config = ConfiguracionCostos.objects.create(
            margen_objetivo=Decimal('35.00'),
            rotacion_objetivo=Decimal('4.00'),
            cobertura_objetivo_dias=30
        )

    if request.method == 'POST':
        try:
            # Actualizar objetivos desde el formulario
            config.margen_objetivo = Decimal(request.POST.get('margen_objetivo', '35.00'))
            config.rotacion_objetivo = Decimal(request.POST.get('rotacion_objetivo', '4.00'))
            config.cobertura_objetivo_dias = int(request.POST.get('cobertura_objetivo_dias', '30'))

            config.save()

            messages.success(
                request,
                '✅ Configuración guardada correctamente. Los cambios se reflejarán en todos los dashboards.'
            )
            return redirect('gestion:configuracion_negocio')

        except Exception as e:
            messages.error(request, f'Error al guardar configuración: {str(e)}')

    context = {
        'config': config
    }

    return render(request, 'gestion/configuracion_negocio.html', context)


# ==================== ⚖️ VISTAS DE AJUSTES DE INVENTARIO ====================

@login_required
def lista_ajustes(request):
    """Vista para listar todos los ajustes de inventario (productos y materias primas)."""
    ajustes = AjusteInventario.objects.select_related(
        'producto', 'materia_prima', 'usuario'
    ).all()

    # Filtros opcionales
    tipo_filtro = request.GET.get('tipo')
    item_tipo_filtro = request.GET.get('item_tipo')  # 'producto' o 'materia_prima'

    if tipo_filtro:
        ajustes = ajustes.filter(tipo=tipo_filtro)

    if item_tipo_filtro == 'producto':
        ajustes = ajustes.filter(producto__isnull=False)
    elif item_tipo_filtro == 'materia_prima':
        ajustes = ajustes.filter(materia_prima__isnull=False)

    context = {
        'ajustes': ajustes,
        'tipo_filtro': tipo_filtro,
        'item_tipo_filtro': item_tipo_filtro,
        'tipos_ajuste': AjusteInventario.TIPO_CHOICES,
    }

    return render(request, 'modules/ajustes/lista.html', context)


@login_required
def crear_ajuste_producto(request, producto_id=None):
    """Vista para crear un ajuste de stock de producto."""
    producto = None
    if producto_id:
        producto = get_object_or_404(Producto, pk=producto_id)

    if request.method == 'POST':
        form = AjusteProductoForm(request.POST)
        if form.is_valid():
            ajuste = form.save(commit=False)
            ajuste.usuario = request.user
            ajuste.save()

            # IMPORTANTE: Actualizar el stock del producto
            producto = ajuste.producto
            producto.stock = ajuste.stock_nuevo
            producto.save()

            messages.success(
                request,
                f'✅ Ajuste registrado: {ajuste.item_nombre} '
                f'{ajuste.stock_anterior}→{ajuste.stock_nuevo} '
                f'({"+"+str(ajuste.diferencia) if ajuste.diferencia > 0 else str(ajuste.diferencia)})'
            )

            # Redirigir al detalle del producto si venimos de ahí
            if producto_id:
                return redirect('gestion:detalle_producto', pk=producto_id)
            return redirect('gestion:lista_ajustes')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        # Si viene producto_id, pasarlo al form
        if producto_id:
            form = AjusteProductoForm(producto_id=producto_id)
        else:
            form = AjusteProductoForm()

    context = {
        'form': form,
        'producto': producto,
        'title': f'Ajustar Stock - {producto.nombre}' if producto else 'Ajustar Stock de Producto',
    }

    return render(request, 'modules/ajustes/form_producto.html', context)


@login_required
def crear_ajuste_materia_prima(request, mp_id=None):
    """Vista para crear un ajuste de stock de materia prima."""
    materia_prima = None
    if mp_id:
        materia_prima = get_object_or_404(MateriaPrima, pk=mp_id)

    if request.method == 'POST':
        form = AjusteMateriaPrimaForm(request.POST)
        if form.is_valid():
            ajuste = form.save(commit=False)
            ajuste.usuario = request.user
            ajuste.save()

            # IMPORTANTE: Actualizar el stock de la materia prima
            mp = ajuste.materia_prima
            mp.stock_actual = ajuste.stock_nuevo
            mp.save()

            messages.success(
                request,
                f'✅ Ajuste registrado: {ajuste.item_nombre} '
                f'{ajuste.stock_anterior}→{ajuste.stock_nuevo} '
                f'({"+"+str(ajuste.diferencia) if ajuste.diferencia > 0 else str(ajuste.diferencia)})'
            )

            # Redirigir al detalle de MP si venimos de ahí
            if mp_id:
                return redirect('gestion:detalle_materia_prima', pk=mp_id)
            return redirect('gestion:lista_ajustes')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        # Si viene mp_id, pasarlo al form
        if mp_id:
            form = AjusteMateriaPrimaForm(materia_prima_id=mp_id)
        else:
            form = AjusteMateriaPrimaForm()

    context = {
        'form': form,
        'materia_prima': materia_prima,
        'title': f'Ajustar Stock - {materia_prima.nombre}' if materia_prima else 'Ajustar Stock de Materia Prima',
    }

    return render(request, 'modules/ajustes/form_materia_prima.html', context)


@login_required
def detalle_ajuste(request, pk):
    """Vista para ver el detalle de un ajuste específico."""
    ajuste = get_object_or_404(
        AjusteInventario.objects.select_related('producto', 'materia_prima', 'usuario'),
        pk=pk
    )

    context = {
        'ajuste': ajuste,
    }

    return render(request, 'modules/ajustes/detalle.html', context)
